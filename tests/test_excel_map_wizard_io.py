from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from vasoanalyzer.ui.dialogs import excel_map_wizard


def test_load_workbook_preserve_uses_keep_vba_for_xlsm(monkeypatch, tmp_path) -> None:
    calls: list[tuple[str, bool, bool]] = []
    sentinel = object()

    def fake_load_workbook(path: str, *, keep_vba: bool, data_only: bool):
        calls.append((path, keep_vba, data_only))
        return sentinel

    monkeypatch.setattr(excel_map_wizard, "load_workbook", fake_load_workbook)

    path = tmp_path / "template.xlsm"
    path.write_text("", encoding="utf-8")

    result = excel_map_wizard.load_workbook_preserve(str(path))

    assert result is sentinel
    assert calls == [(str(path), True, False)]


def test_load_workbook_preserve_disables_keep_vba_for_xlsx(monkeypatch, tmp_path) -> None:
    calls: list[tuple[str, bool, bool]] = []
    sentinel = object()

    def fake_load_workbook(path: str, *, keep_vba: bool, data_only: bool):
        calls.append((path, keep_vba, data_only))
        return sentinel

    monkeypatch.setattr(excel_map_wizard, "load_workbook", fake_load_workbook)

    path = tmp_path / "template.xlsx"
    path.write_text("", encoding="utf-8")

    result = excel_map_wizard.load_workbook_preserve(str(path))

    assert result is sentinel
    assert calls == [(str(path), False, False)]


def test_save_workbook_writes_temp_and_replaces_target(monkeypatch, tmp_path) -> None:
    target = tmp_path / "mapped.xlsm"
    target.write_text("old", encoding="utf-8")
    wb = Workbook()
    calls: dict[str, object] = {}

    def fake_save(self: Workbook, filename: str) -> None:
        calls["save_path"] = filename
        Path(filename).write_text("new", encoding="utf-8")

    def fake_replace(src: str, dst: str) -> None:
        calls["replace"] = (src, dst)
        Path(dst).write_text(Path(src).read_text(encoding="utf-8"), encoding="utf-8")
        Path(src).unlink()

    monkeypatch.setattr(Workbook, "save", fake_save, raising=True)
    monkeypatch.setattr(excel_map_wizard.os, "replace", fake_replace)

    excel_map_wizard.save_workbook(wb, str(target))

    save_path = Path(calls["save_path"])
    assert save_path.parent == target.parent
    assert save_path.suffix == target.suffix
    assert save_path != target
    assert calls["replace"] == (str(save_path), str(target))
    assert target.read_text(encoding="utf-8") == "new"
    assert not save_path.exists()
