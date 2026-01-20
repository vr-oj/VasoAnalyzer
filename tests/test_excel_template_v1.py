from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from vasoanalyzer.excel.template_v1 import inspect_template, validate_template_or_raise
from vasoanalyzer.excel.writer_v1 import (
    apply_write_plan,
    build_write_plan,
    validate_write_plan_or_raise,
)


def _template_path() -> Path:
    root = Path(__file__).resolve().parents[1]
    return root / "src/vasoanalyzer/resources/templates/VasoAnalyzer_Standard_Template_v1.xlsx"


def _first_block():
    inspection = inspect_template(str(_template_path()))
    return inspection.blocks[0]


def test_inspect_template_detects_signature_and_blocks():
    inspection = inspect_template(str(_template_path()))
    assert inspection.signature_ok is True
    assert inspection.blocks
    block = inspection.blocks[0]
    assert block.metric_col == "Metric"
    assert "Replicate_1" in block.replicate_cols
    assert all(name in block.summary_cols for name in ("MEAN", "SEM", "N"))


def test_build_write_plan_maps_metrics_to_rows(tmp_path):
    block = _first_block()
    export_table = pd.DataFrame(
        [
            {"Metric": "20 mmHg – Max", "Value": 101.5},
            {"Metric": "40 mmHg – Max", "Value": 99.2},
        ]
    )
    output_path = tmp_path / "filled.xlsx"
    plan = build_write_plan(
        str(_template_path()), str(output_path), block, "Replicate_1", export_table
    )
    assert not plan.missing_metrics
    assert len(plan.items) == 2
    assert all(item.target_cell.startswith("B") for item in plan.items)


def test_missing_metrics_fail_fast(tmp_path):
    block = _first_block()
    export_table = pd.DataFrame([{"Metric": "Missing Metric", "Value": 1.0}])
    output_path = tmp_path / "filled.xlsx"
    plan = build_write_plan(
        str(_template_path()), str(output_path), block, "Replicate_1", export_table
    )
    try:
        validate_write_plan_or_raise(plan)
        assert False, "Expected validation to fail for missing metrics."
    except ValueError as exc:
        assert "Missing metrics" in str(exc)


def test_apply_write_plan_writes_values_and_preserves_formulas(tmp_path):
    template_path = _template_path()
    block = _first_block()
    export_table = pd.DataFrame(
        [
            {"Metric": "20 mmHg – Max", "Value": 101.5},
            {"Metric": "60 mmHg – Max", "Value": 88.25},
        ]
    )
    output_path = tmp_path / "filled.xlsx"
    plan = build_write_plan(
        str(template_path), str(output_path), block, "Replicate_1", export_table
    )
    apply_write_plan(plan)

    wb = load_workbook(output_path, data_only=False)
    ws = wb[block.sheet_name]
    table = next(
        table for table in ws.tables.values() if (table.displayName or table.name) == block.table_name
    )
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    col_names = [col.name for col in table.tableColumns]
    metric_col = min_col + col_names.index("Metric")
    replicate_col = min_col + col_names.index("Replicate_1")
    mean_col = min_col + col_names.index("MEAN")

    values_by_metric = {
        ws.cell(row=row, column=metric_col).value: row
        for row in range(min_row + 1, max_row + 1)
        if ws.cell(row=row, column=metric_col).value is not None
    }
    row_20 = values_by_metric["20 mmHg – Max"]
    row_60 = values_by_metric["60 mmHg – Max"]

    assert ws.cell(row=row_20, column=replicate_col).value == 101.5
    assert ws.cell(row=row_60, column=replicate_col).value == 88.25
    assert str(ws.cell(row=row_20, column=mean_col).value).startswith("=")
