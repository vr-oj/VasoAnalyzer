"""Dialog for exporting session data into either a VasoAnalyzer standard template
or any arbitrary lab Excel template (flexible mode).
"""

from __future__ import annotations

import importlib.resources
import shutil
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QColor
from PyQt6.QtWidgets import (
    QCheckBox,
    QComboBox,
    QDialog,
    QFileDialog,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
)

from vasoanalyzer.excel.flexible_writer import (
    apply_flexible_write_plan,
    build_flexible_write_plan,
)
from vasoanalyzer.excel.label_matching import best_match as _find_best_match
from vasoanalyzer.excel.template_metadata import read_template_metadata
from vasoanalyzer.excel.template_v1 import inspect_template, validate_template_or_raise
from vasoanalyzer.excel.writer_v1 import apply_write_plan, build_write_plan
from vasoanalyzer.export.generator import build_export_table, events_from_rows
from vasoanalyzer.export.profiles import EVENT_TABLE_ROW_PER_EVENT, PRESSURE_CURVE_STANDARD


class ExcelTemplateExportDialog(QDialog):
    """Export dialog that auto-detects Standard vs Flexible template mode."""

    def __init__(self, parent, *, event_rows: list[tuple]) -> None:
        super().__init__(parent)
        self.setWindowTitle("Export to Excel Template")
        self.setMinimumSize(740, 560)

        self._event_rows = list(event_rows or [])
        self._mode = "standard"  # "standard" | "flexible"

        # Standard-mode state
        self._inspection = None
        self._blocks = []
        self._current_plan = None

        # Flexible-mode state
        self._flexible_metadata = None
        self._flexible_plan = None
        self._flex_wb = None  # cached workbook for column scanning

        layout = QVBoxLayout(self)

        # --- File group ----------------------------------------------------------
        file_group = QGroupBox("Template Workbook")
        file_layout = QGridLayout(file_group)
        file_layout.addWidget(QLabel("Workbook:"), 0, 0)
        self.template_path_edit = QLineEdit()
        self.template_browse_btn = QPushButton("Browse…")
        self._get_template_btn = QPushButton("Get Standard Template…")
        self._get_template_btn.setToolTip(
            "Save a copy of the VasoAnalyzer Standard Template to your computer"
        )
        file_layout.addWidget(self.template_path_edit, 0, 1)
        file_layout.addWidget(self.template_browse_btn, 0, 2)
        file_layout.addWidget(self._get_template_btn, 0, 3)

        self._mode_label = QLabel("")
        self._mode_label.setStyleSheet("color: #6B7280; font-style: italic;")
        file_layout.addWidget(self._mode_label, 1, 1, 1, 3)
        layout.addWidget(file_group)

        # --- Standard mode options -----------------------------------------------
        self.options_group = QGroupBox("Export Options")
        options_layout = QGridLayout(self.options_group)
        options_layout.addWidget(QLabel("Block:"), 0, 0)
        self.block_combo = QComboBox()
        options_layout.addWidget(self.block_combo, 0, 1, 1, 2)
        options_layout.addWidget(QLabel("Replicate:"), 1, 0)
        self.replicate_combo = QComboBox()
        options_layout.addWidget(self.replicate_combo, 1, 1, 1, 2)
        options_layout.addWidget(QLabel("Profile:"), 2, 0)
        self.profile_combo = QComboBox()
        self.profile_combo.addItem(PRESSURE_CURVE_STANDARD.display_name, PRESSURE_CURVE_STANDARD)
        options_layout.addWidget(self.profile_combo, 2, 1, 1, 2)
        self.skip_missing_cb = QCheckBox("Skip missing metrics (not recommended)")
        options_layout.addWidget(self.skip_missing_cb, 3, 0, 1, 3)
        layout.addWidget(self.options_group)

        # --- Flexible mode options -----------------------------------------------
        self._flexible_group = QGroupBox("Template Mapping")
        flex_layout = QVBoxLayout(self._flexible_group)

        # Auto-detect info banner (shown when source = "inferred")
        self._infer_banner = QLabel(
            "Structure was auto-detected — verify the row mapping below before writing."
        )
        self._infer_banner.setWordWrap(True)
        self._infer_banner.setStyleSheet(
            "background: #FEF3C7; color: #92400E; padding: 6px 8px;"
            "border: 1px solid #F59E0B; border-radius: 4px;"
        )
        self._infer_banner.setVisible(False)
        flex_layout.addWidget(self._infer_banner)

        # Sheet + source row
        flex_top = QGridLayout()
        flex_top.addWidget(QLabel("Sheet:"), 0, 0)
        self._sheet_combo = QComboBox()
        self._sheet_combo.setMinimumWidth(160)
        flex_top.addWidget(self._sheet_combo, 0, 1)
        self._source_label = QLabel("")
        self._source_label.setStyleSheet("color: #6B7280; font-style: italic;")
        flex_top.addWidget(self._source_label, 0, 2)
        flex_top.setColumnStretch(3, 1)

        # Column to write into
        flex_top.addWidget(QLabel("Write into column:"), 1, 0)
        self._col_combo = QComboBox()
        self._col_combo.setMinimumWidth(260)
        flex_top.addWidget(self._col_combo, 1, 1, 1, 2)
        flex_layout.addLayout(flex_top)

        # Row mapping table
        flex_layout.addWidget(QLabel("Row Mapping  (assign each template row to a session event):"))
        self._mapping_table = QTableWidget(0, 2)
        self._mapping_table.setHorizontalHeaderLabels(["Template Row Label", "Session Event"])
        self._mapping_table.horizontalHeader().setStretchLastSection(True)
        self._mapping_table.horizontalHeader().setDefaultSectionSize(260)
        self._mapping_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._mapping_table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        flex_layout.addWidget(self._mapping_table, 1)

        self._flexible_group.setVisible(False)
        layout.addWidget(self._flexible_group)

        # --- Output path ---------------------------------------------------------
        output_group = QGroupBox("Output")
        output_layout = QGridLayout(output_group)
        output_layout.addWidget(QLabel("Save as:"), 0, 0)
        self.output_path_edit = QLineEdit()
        self.output_browse_btn = QPushButton("Browse…")
        output_layout.addWidget(self.output_path_edit, 0, 1)
        output_layout.addWidget(self.output_browse_btn, 0, 2)
        layout.addWidget(output_group)

        # --- Preview table -------------------------------------------------------
        preview_group = QGroupBox("Write Plan Preview")
        preview_layout = QVBoxLayout(preview_group)
        self.preview_table = QTableWidget(0, 3)
        self.preview_table.setHorizontalHeaderLabels(["Row / Metric", "Value", "Target Cell"])
        self.preview_table.horizontalHeader().setStretchLastSection(True)
        self.preview_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        preview_layout.addWidget(self.preview_table)
        self.status_label = QLabel("")
        self.status_label.setWordWrap(True)
        self.status_label.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop)
        preview_layout.addWidget(self.status_label)
        layout.addWidget(preview_group, 1)

        # --- Button row ----------------------------------------------------------
        btn_row = QHBoxLayout()
        self.preview_btn = QPushButton("Preview")
        self.write_btn = QPushButton("Write")
        self.write_btn.setEnabled(False)
        cancel_btn = QPushButton("Close")
        btn_row.addStretch()
        btn_row.addWidget(self.preview_btn)
        btn_row.addWidget(self.write_btn)
        btn_row.addWidget(cancel_btn)
        layout.addLayout(btn_row)

        # --- Connections ---------------------------------------------------------
        self.template_browse_btn.clicked.connect(self._choose_template)
        self.output_browse_btn.clicked.connect(self._choose_output)
        self.block_combo.currentIndexChanged.connect(self._refresh_replicates)
        self.preview_btn.clicked.connect(self._preview_plan)
        self.write_btn.clicked.connect(self._write_plan)
        cancel_btn.clicked.connect(self.reject)
        self._get_template_btn.clicked.connect(self._export_standard_template)
        self._sheet_combo.currentTextChanged.connect(self._on_sheet_changed)

    # -------------------------------------------------------------------------
    # File selection
    # -------------------------------------------------------------------------

    def _choose_template(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Excel template",
            "",
            "Excel Files (*.xlsx *.xlsm);;All Files (*)",
        )
        if not path:
            return
        self.template_path_edit.setText(path)
        self._load_template(path)

    def _choose_output(self) -> None:
        start = self.output_path_edit.text().strip()
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Save output workbook",
            start or "",
            "Excel Files (*.xlsx);;All Files (*)",
        )
        if path:
            self.output_path_edit.setText(path)

    def _export_standard_template(self) -> None:
        dest, _ = QFileDialog.getSaveFileName(
            self,
            "Save Standard Template",
            "VasoAnalyzer_Standard_Template_v1.xlsx",
            "Excel Files (*.xlsx);;All Files (*)",
        )
        if not dest:
            return
        try:
            pkg = importlib.resources.files("vasoanalyzer.resources.templates")
            src = pkg.joinpath("VasoAnalyzer_Standard_Template_v1.xlsx")
            with importlib.resources.as_file(src) as src_path:
                shutil.copy2(src_path, dest)
            QMessageBox.information(
                self,
                "Template Saved",
                f"Standard template saved to:\n{dest}",
            )
        except Exception as exc:
            QMessageBox.critical(self, "Error", f"Could not save template:\n{exc}")

    # -------------------------------------------------------------------------
    # Template loading and mode detection
    # -------------------------------------------------------------------------

    def _load_template(self, path: str) -> None:
        """Detect Standard vs Flexible mode and update UI accordingly."""
        try:
            inspection = inspect_template(path)
        except Exception as exc:
            QMessageBox.critical(self, "Error reading template", str(exc))
            return

        if inspection.signature_ok and inspection.blocks:
            self._switch_to_standard(path, inspection)
        else:
            self._switch_to_flexible(path)

    def _switch_to_standard(self, path: str, inspection) -> None:
        self._mode = "standard"
        self._mode_label.setText("Standard VasoAnalyzer Template detected")
        self._mode_label.setStyleSheet("color: #059669; font-style: italic;")
        self.options_group.setVisible(True)
        self._flexible_group.setVisible(False)

        try:
            validate_template_or_raise(inspection)
        except Exception as exc:
            QMessageBox.critical(self, "Invalid Template", str(exc))
            self._inspection = None
            self._blocks = []
            self.block_combo.clear()
            self.replicate_combo.clear()
            self.write_btn.setEnabled(False)
            return

        self._inspection = inspection
        self._blocks = inspection.blocks
        self.block_combo.clear()
        for block in self._blocks:
            label = block.title or block.block_id
            self.block_combo.addItem(f"{label} ({block.block_id})", block.block_id)
        self._refresh_replicates()
        self._set_default_output_path(path)
        self.status_label.setText("\n".join(inspection.warnings) if inspection.warnings else "")

    def _switch_to_flexible(self, path: str) -> None:
        self._mode = "flexible"
        self._mode_label.setText("Custom template detected — configure mapping below")
        self._mode_label.setStyleSheet("color: #D97706; font-style: italic;")
        self.options_group.setVisible(False)
        self._flexible_group.setVisible(True)

        try:
            wb = load_workbook(path, data_only=False)
        except Exception as exc:
            QMessageBox.critical(self, "Error reading workbook", str(exc))
            return

        self._flex_wb = wb
        self._sheet_combo.blockSignals(True)
        self._sheet_combo.clear()
        for name in wb.sheetnames:
            ws = wb[name]
            if ws.sheet_state != "hidden":
                self._sheet_combo.addItem(name)
        self._sheet_combo.blockSignals(False)

        self._set_default_output_path(path)
        self._reload_flexible_metadata(path)

    def _on_sheet_changed(self, sheet_name: str) -> None:
        path = self.template_path_edit.text().strip()
        if path and sheet_name:
            self._reload_flexible_metadata(path)

    def _reload_flexible_metadata(self, path: str) -> None:
        sheet_name = self._sheet_combo.currentText()
        if not sheet_name:
            return

        try:
            wb = load_workbook(path, data_only=False)
            wb.active = wb[sheet_name]
            metadata = read_template_metadata(path, wb)
        except Exception as exc:
            self.status_label.setText(f"Could not read template structure: {exc}")
            return

        self._flexible_metadata = metadata
        if metadata is None:
            self.status_label.setText(
                "Could not detect template structure. "
                "Ensure the sheet has labels in column A and data in columns B onwards."
            )
            return

        self._populate_flexible_ui(metadata)

    def _populate_flexible_ui(self, metadata) -> None:
        source_labels = {
            "vba_metadata": "VasoAnalyzer metadata",
            "named_ranges": "Named ranges",
            "inferred": "Auto-detected",
            "sheet_specific": "Sheet metadata",
        }
        src_text = source_labels.get(metadata.source, metadata.source)
        self._source_label.setText(f"Detection: {src_text}")
        self._infer_banner.setVisible(metadata.source == "inferred")

        # Column picker — only non-formula columns
        self._col_combo.clear()
        for dc in metadata.date_columns:
            parts = [f"Column {dc.letter}"]
            if dc.value:
                parts.append(dc.value)
            if dc.empty_slots:
                parts.append(f"{dc.empty_slots} empty slot(s)")
            self._col_combo.addItem(" — ".join(parts), dc.column)

        # Default to column with most empty slots (best candidate for new data)
        if metadata.date_columns:
            max_empty = max(dc.empty_slots for dc in metadata.date_columns)
            for i, dc in enumerate(metadata.date_columns):
                if dc.empty_slots == max_empty:
                    self._col_combo.setCurrentIndex(i)
                    break
            # Offer "add new column" after existing columns
            next_col = max(dc.column for dc in metadata.date_columns) + 1
            from openpyxl.utils import get_column_letter
            self._col_combo.addItem(
                f"+ New column ({get_column_letter(next_col)})", next_col
            )

        # Row mapping table
        session_labels = self._get_session_event_labels()
        choices = ["(skip)"] + session_labels

        self._mapping_table.setRowCount(0)
        for event_row in metadata.event_rows:
            row_idx = self._mapping_table.rowCount()
            self._mapping_table.insertRow(row_idx)

            if event_row.is_header:
                item = QTableWidgetItem(f"— {event_row.label} —")
                item.setFlags(Qt.ItemFlag.ItemIsEnabled)
                item.setForeground(QColor("#9CA3AF"))
                self._mapping_table.setItem(row_idx, 0, item)
                blank = QTableWidgetItem("")
                blank.setFlags(Qt.ItemFlag.ItemIsEnabled)
                blank.setForeground(QColor("#9CA3AF"))
                self._mapping_table.setItem(row_idx, 1, blank)
            else:
                label_item = QTableWidgetItem(event_row.label)
                label_item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
                self._mapping_table.setItem(row_idx, 0, label_item)

                combo = QComboBox()
                combo.addItems(choices)
                best = _find_best_match(event_row.label, session_labels)
                if best:
                    idx = combo.findText(best)
                    if idx >= 0:
                        combo.setCurrentIndex(idx)
                self._mapping_table.setCellWidget(row_idx, 1, combo)

        self.status_label.setText("")

    # -------------------------------------------------------------------------
    # Helpers
    # -------------------------------------------------------------------------

    def _get_session_event_labels(self) -> list[str]:
        """Return unique event labels from the current session."""
        try:
            events = events_from_rows(self._event_rows)
            table = build_export_table(EVENT_TABLE_ROW_PER_EVENT, events)
            seen: list[str] = []
            for row in table.rows:
                if len(row) >= 2:
                    label = str(row[1])
                    if label not in seen:
                        seen.append(label)
            return seen
        except Exception:
            return []

    def _get_session_event_values(self) -> dict[str, float]:
        """Return a mapping of event label → numeric value for the current session."""
        try:
            events = events_from_rows(self._event_rows)
            table = build_export_table(EVENT_TABLE_ROW_PER_EVENT, events)
            result: dict[str, float] = {}
            for row in table.rows:
                if len(row) >= 3:
                    label = str(row[1])
                    try:
                        result[label] = float(row[2])
                    except (TypeError, ValueError):
                        pass
            return result
        except Exception:
            return {}

    def _get_flexible_mapping(self) -> dict[str, str | None]:
        """Read the current state of the row-mapping table."""
        mapping: dict[str, str | None] = {}
        for row_idx in range(self._mapping_table.rowCount()):
            label_item = self._mapping_table.item(row_idx, 0)
            widget = self._mapping_table.cellWidget(row_idx, 1)
            if label_item is None or widget is None:
                continue
            label = label_item.text()
            if label.startswith("—") and label.endswith("—"):
                continue  # section header row
            chosen = widget.currentText()
            mapping[label] = None if chosen == "(skip)" else chosen
        return mapping

    def _set_default_output_path(self, template_path: str) -> None:
        path = Path(template_path)
        if not path.name:
            return
        output = path.with_name(f"{path.stem}_filled{path.suffix}")
        self.output_path_edit.setText(str(output))

    def _refresh_replicates(self) -> None:
        self.replicate_combo.clear()
        block = self._selected_block()
        if block is None:
            return
        ordered = sorted(
            block.replicate_cols,
            key=lambda name: int(name.split("_")[-1]) if name.split("_")[-1].isdigit() else 0,
        )
        self.replicate_combo.addItems(ordered)

    def _selected_block(self):
        block_id = self.block_combo.currentData()
        if not block_id:
            return None
        for block in self._blocks:
            if block.block_id == block_id:
                return block
        return None

    def _profile_to_dataframe(self) -> pd.DataFrame:
        profile = self.profile_combo.currentData()
        events = events_from_rows(self._event_rows)
        table = build_export_table(profile, events)
        rows = [(row[0], row[1]) for row in table.rows if len(row) >= 2]
        return pd.DataFrame(rows, columns=["Metric", "Value"])

    # -------------------------------------------------------------------------
    # Preview & Write — dispatches to standard or flexible
    # -------------------------------------------------------------------------

    def _preview_plan(self) -> None:
        if self._mode == "flexible":
            self._preview_flexible_plan()
        else:
            self._preview_standard_plan()

    def _write_plan(self) -> None:
        if self._mode == "flexible":
            self._write_flexible_plan()
        else:
            self._write_standard_plan()

    # Standard mode -----------------------------------------------------------

    def _preview_standard_plan(self) -> None:
        template_path = self.template_path_edit.text().strip()
        output_path = self.output_path_edit.text().strip()
        block = self._selected_block()
        replicate = self.replicate_combo.currentText()

        self.preview_table.setRowCount(0)
        self.write_btn.setEnabled(False)

        if not (template_path and output_path and block and replicate):
            self.status_label.setText("Select a template, block, replicate, and output path.")
            return

        df = self._profile_to_dataframe()
        try:
            plan = build_write_plan(template_path, output_path, block, replicate, df)
        except Exception as exc:
            self.status_label.setText(str(exc))
            return

        original_missing = list(plan.missing_metrics)
        if plan.missing_metrics and self.skip_missing_cb.isChecked():
            filtered = df[~df["Metric"].isin(plan.missing_metrics)]
            plan = build_write_plan(template_path, output_path, block, replicate, filtered)
            if plan.missing_metrics:
                self.status_label.setText("Missing metrics remain after filtering.")
                return

        self._current_plan = plan
        for item in plan.items:
            row_idx = self.preview_table.rowCount()
            self.preview_table.insertRow(row_idx)
            self.preview_table.setItem(row_idx, 0, QTableWidgetItem(item.metric))
            self.preview_table.setItem(row_idx, 1, QTableWidgetItem(f"{item.value:.2f}"))
            self.preview_table.setItem(row_idx, 2, QTableWidgetItem(item.target_cell))

        messages = []
        if original_missing:
            messages.append("Missing metrics: " + ", ".join(original_missing))
        if plan.warnings:
            messages.extend(plan.warnings)
        self.status_label.setText("\n".join(messages))
        self.write_btn.setEnabled(not original_missing or self.skip_missing_cb.isChecked())

    def _write_standard_plan(self) -> None:
        if self._current_plan is None:
            self._preview_standard_plan()
            if self._current_plan is None:
                return
        try:
            apply_write_plan(self._current_plan)
        except Exception as exc:
            QMessageBox.critical(self, "Export Failed", str(exc))
            return
        QMessageBox.information(
            self,
            "Export Complete",
            f"Workbook saved to:\n{self._current_plan.output_path}",
        )
        self.accept()

    # Flexible mode -----------------------------------------------------------

    def _preview_flexible_plan(self) -> None:
        template_path = self.template_path_edit.text().strip()
        output_path = self.output_path_edit.text().strip()
        sheet_name = self._sheet_combo.currentText()
        target_col = self._col_combo.currentData()

        self.preview_table.setRowCount(0)
        self.write_btn.setEnabled(False)

        if not (template_path and output_path and sheet_name and target_col):
            self.status_label.setText("Select a template, sheet, target column, and output path.")
            return

        if self._flexible_metadata is None:
            self.status_label.setText("Template structure could not be detected.")
            return

        label_to_event = self._get_flexible_mapping()
        event_values = self._get_session_event_values()

        try:
            plan = build_flexible_write_plan(
                label_to_event=label_to_event,
                event_label_to_value=event_values,
                metadata=self._flexible_metadata,
                target_col=target_col,
                template_path=template_path,
                output_path=output_path,
                sheet_name=sheet_name,
            )
        except Exception as exc:
            self.status_label.setText(str(exc))
            return

        self._flexible_plan = plan

        from openpyxl.utils import get_column_letter

        for row_idx_write, col_idx_write, value in plan.writes:
            row_idx = self.preview_table.rowCount()
            self.preview_table.insertRow(row_idx)
            # Find the label for this row
            label = ""
            if self._flexible_metadata:
                for er in self._flexible_metadata.event_rows:
                    if er.row == row_idx_write:
                        label = er.label
                        break
            cell_ref = f"{get_column_letter(col_idx_write)}{row_idx_write}"
            self.preview_table.setItem(row_idx, 0, QTableWidgetItem(label))
            val_str = f"{value:.4g}" if value is not None else "—"
            self.preview_table.setItem(row_idx, 1, QTableWidgetItem(val_str))
            self.preview_table.setItem(row_idx, 2, QTableWidgetItem(cell_ref))

        self.status_label.setText("\n".join(plan.warnings) if plan.warnings else "")
        self.write_btn.setEnabled(bool(plan.writes))

    def _write_flexible_plan(self) -> None:
        if self._flexible_plan is None:
            self._preview_flexible_plan()
            if self._flexible_plan is None:
                return
        try:
            apply_flexible_write_plan(self._flexible_plan)
        except Exception as exc:
            QMessageBox.critical(self, "Export Failed", str(exc))
            return
        QMessageBox.information(
            self,
            "Export Complete",
            f"Workbook saved to:\n{self._flexible_plan.output_path}",
        )
        self.accept()
