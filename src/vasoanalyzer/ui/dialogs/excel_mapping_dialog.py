# VasoAnalyzer
# Copyright © 2025 Osvaldo J. Vega Rodríguez
# Licensed under CC BY-NC-SA 4.0 International
# http://creativecommons.org/licenses/by-nc-sa/4.0/

import logging
import os
import subprocess
import sys
import time

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from PyQt5.QtGui import QColor, QPalette
from PyQt5.QtWidgets import (
    QComboBox,
    QDialog,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
)

from vasoanalyzer.ui.theme import CURRENT_THEME

log = logging.getLogger(__name__)


class ExcelMappingDialog(QDialog):
    def __init__(self, parent, event_data):
        super().__init__(parent)
        self.setWindowTitle("Map Events to Excel")
        bg = CURRENT_THEME["window_bg"]
        text = CURRENT_THEME["text"]
        border = CURRENT_THEME.get("grid_color", text)
        button_bg = CURRENT_THEME.get("button_bg", bg)
        button_hover = CURRENT_THEME.get("button_hover_bg", CURRENT_THEME.get("selection_bg", bg))
        selection_bg = CURRENT_THEME.get("selection_bg", button_hover)
        table_bg = CURRENT_THEME.get("table_bg", bg)
        alternate_bg = CURRENT_THEME.get("alternate_bg", table_bg)
        self.setStyleSheet(
            f"""
            QLabel, QTableWidget, QComboBox, QPushButton {{
                color: {text};
                font-family: Arial;
                font-size: 13px;
            }}
            QComboBox {{
                background-color: {button_bg};
                border: 1px solid {border};
                border-radius: 4px;
                padding: 4px;
            }}
            QComboBox QAbstractItemView {{
                background-color: {bg};
                selection-background-color: {selection_bg};
                color: {text};
            }}
            QPushButton {{
                background-color: {button_bg};
                border: 1px solid {border};
                border-radius: 6px;
                padding: 6px 12px;
                color: {text};
            }}
            QPushButton:hover {{
                background-color: {button_hover};
            }}
            QTableWidget {{
                background-color: {table_bg};
                alternate-background-color: {alternate_bg};
            }}
        """
        )
        self.event_data = event_data
        self.excel_path = None
        self.wb = None
        self.ws = None
        self.current_row = 3
        self.selected_column = None
        self.history = []

        self.setMinimumWidth(460)
        self.layout = QVBoxLayout(self)
        self.layout.setSpacing(12)

        self.instructions = QLabel("<b>Step 1:</b> Select Excel file")
        self.layout.addWidget(self.instructions)

        self.load_button = QPushButton("Load Excel Template")
        self.load_button.clicked.connect(self.load_excel)
        self.layout.addWidget(self.load_button)

        self.excel_filename_label = QLabel("")
        self.layout.addWidget(self.excel_filename_label)

        self.layout.addWidget(QLabel("<b>Step 2:</b> Select column to populate:"))
        self.column_selector = QComboBox()
        self.column_selector.addItems([chr(i) for i in range(66, 91)])
        self.column_selector.setEnabled(False)
        self.column_selector.currentTextChanged.connect(self.update_preview_table)
        self.layout.addWidget(self.column_selector)

        self.layout.addSpacing(6)
        self.cell_label = QLabel("Next Excel Cell: N/A")
        self.layout.addWidget(self.cell_label)

        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)
        self.layout.addWidget(line)

        self.event_table = QTableWidget()
        self.event_table.setColumnCount(4)
        self.event_table.setHorizontalHeaderLabels(
            ["EventLabel", "Time (s)", "ID (\u00b5m)", "Frame"]
        )
        self.event_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.event_table.cellClicked.connect(self.map_event_to_excel)
        self.event_table.setMinimumWidth(420)
        header = self.event_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Interactive)
        header.setMinimumSectionSize(80)
        header.setStretchLastSection(True)
        header.setDefaultSectionSize(100)

        vheader = self.event_table.verticalHeader()
        vheader.setSectionResizeMode(QHeaderView.Fixed)
        vheader.setDefaultSectionSize(24)
        self.layout.addWidget(self.event_table)
        self.populate_event_table()

        # Preview table for selected Excel column
        self.preview_table = QTableWidget()
        self.preview_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.preview_table.setMinimumWidth(420)
        pheader = self.preview_table.horizontalHeader()
        pheader.setSectionResizeMode(QHeaderView.Interactive)
        pheader.setMinimumSectionSize(80)
        pheader.setStretchLastSection(True)
        pheader.setDefaultSectionSize(100)

        pvheader = self.preview_table.verticalHeader()
        pvheader.setSectionResizeMode(QHeaderView.Fixed)
        pvheader.setDefaultSectionSize(24)
        self.layout.addWidget(self.preview_table)

        # Apply theme to tables
        self._apply_table_theme(self.event_table)
        self._apply_template_table_theme(self.preview_table)

        self.button_layout = QHBoxLayout()
        self.button_layout.addStretch()
        self.skip_button = QPushButton("Skip")
        self.skip_button.clicked.connect(self.skip_cell)
        self.undo_button = QPushButton("Undo Last")
        self.undo_button.clicked.connect(self.undo_last)
        self.done_button = QPushButton("Done")
        self.done_button.clicked.connect(self.finish_and_save)
        self.button_layout.addWidget(self.skip_button)
        self.button_layout.addWidget(self.undo_button)
        self.button_layout.addWidget(self.done_button)
        self.layout.addLayout(self.button_layout)

    def _apply_table_theme(self, table) -> None:
        """Apply theme-aware palette to a table widget."""
        palette = table.palette()

        table_bg = CURRENT_THEME.get("table_bg", "#020617")
        alt_bg = CURRENT_THEME.get("alternate_bg", table_bg)
        text = CURRENT_THEME.get("table_text", CURRENT_THEME.get("text", "#FFFFFF"))
        highlight = CURRENT_THEME.get("selection_bg", "#1D4ED8")
        grid = CURRENT_THEME.get("grid_color", "#374151")

        palette.setColor(QPalette.Base, QColor(table_bg))
        palette.setColor(QPalette.AlternateBase, QColor(alt_bg))
        palette.setColor(QPalette.Text, QColor(text))
        palette.setColor(QPalette.WindowText, QColor(text))
        palette.setColor(QPalette.Highlight, QColor(highlight))
        palette.setColor(QPalette.HighlightedText, QColor(text))

        table.setPalette(palette)
        table.setAlternatingRowColors(True)
        table.setStyleSheet(
            f"""
            QTableView {{
                gridline-color: {grid};
                background: {table_bg};
                color: {text};
                selection-background-color: {highlight};
            }}
            QTableWidget {{
                background: {table_bg};
                alternate-background-color: {alt_bg};
                color: {text};
                selection-background-color: {highlight};
            }}
            QHeaderView::section {{
                background: {table_bg};
                color: {text};
            }}
        """
        )

    def _apply_template_table_theme(self, table) -> None:
        """Apply theme-aware palette to the Template Preview table."""
        palette = table.palette()

        table_bg = CURRENT_THEME.get("table_bg", "#020617")
        alt_bg = CURRENT_THEME.get("alternate_bg", table_bg)
        text = CURRENT_THEME.get("table_text", CURRENT_THEME.get("text", "#FFFFFF"))
        highlight = CURRENT_THEME.get("selection_bg", "#1D4ED8")
        grid = CURRENT_THEME.get("grid_color", "#374151")

        palette.setColor(QPalette.Base, QColor(table_bg))
        palette.setColor(QPalette.AlternateBase, QColor(alt_bg))
        palette.setColor(QPalette.Text, QColor(text))
        palette.setColor(QPalette.WindowText, QColor(text))
        palette.setColor(QPalette.Highlight, QColor(highlight))
        palette.setColor(QPalette.HighlightedText, QColor(text))

        table.setPalette(palette)
        table.setAlternatingRowColors(True)
        table.setStyleSheet(
            f"""
            QTableView {{
                gridline-color: {grid};
                background: {table_bg};
                color: {text};
                selection-background-color: {highlight};
            }}
            QTableWidget {{
                background: {table_bg};
                alternate-background-color: {alt_bg};
                color: {text};
                selection-background-color: {highlight};
            }}
            QHeaderView::section {{
                background: {table_bg};
                color: {text};
            }}
        """
        )

    def populate_event_table(self):
        self.event_table.setRowCount(len(self.event_data))
        for i, event in enumerate(self.event_data):
            if isinstance(event, dict):
                label = event.get("EventLabel", "")
                time = event.get("Time (s)", "")
                id_val = event.get("ID (µm)", "")
                frame = event.get("Frame", "")
            else:
                label, time, id_val = event
            self.event_table.setItem(i, 0, QTableWidgetItem(str(label)))
            self.event_table.setItem(i, 1, QTableWidgetItem(str(time)))
            self.event_table.setItem(i, 2, QTableWidgetItem(str(id_val)))
            self.event_table.setItem(i, 3, QTableWidgetItem(str(frame)))
        self.event_table.resizeColumnsToContents()

        self.event_table.setAlternatingRowColors(True)
        self.event_table.setStyleSheet(
            f"""
            QTableWidget {{
                alternate-background-color: {CURRENT_THEME.get("alternate_bg", "#F5F5F5")};
                background-color: {CURRENT_THEME.get("table_bg", "#FFFFFF")};
            }}
        """
        )

    def load_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select Excel File", "", "Excel Files (*.xlsx)")
        if path:
            try:
                self.wb = load_workbook(path)
                self.ws = self.wb.active
                self.excel_path = path
                # Ensure the first column remains visible when scrolling
                if not self.ws.freeze_panes:
                    self.ws.freeze_panes = "B1"
                    self.wb.save(path)

                self.column_selector.setEnabled(True)
                self.instructions.setText(
                    "<b>Step 2:</b> File loaded. Now select column and assign values."
                )
                self.update_cell_label()
                self.update_preview_table()
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to load Excel file:\n{e}")
        self.excel_filename_label.setText(f"<i>Loaded:</i> {os.path.basename(path)}")

    def get_current_cell(self):
        col_letter = self.column_selector.currentText()
        return f"{col_letter}{self.current_row}" if col_letter else None

    def update_cell_label(self):
        col_letter = self.column_selector.currentText()
        cell = f"{col_letter}{self.current_row}" if col_letter else "N/A"
        description = ""
        if self.ws:
            try:
                desc_value = str(self.ws[f"A{self.current_row}"].value)
                if desc_value:
                    description = f" \u2192 <i>{desc_value}</i>"
            except (KeyError, AttributeError, TypeError):
                # Cell doesn't exist or has no value
                pass
        self.cell_label.setText(f"<b>Editing Cell:</b> {cell}{description}")
        self.update_preview_table()

    def update_preview_table(self):
        if not self.ws or not self.column_selector.currentText():
            self.preview_table.setRowCount(0)
            self.preview_table.setColumnCount(0)
            return

        table_bg = CURRENT_THEME.get("table_bg", "#020617")
        alt_bg = CURRENT_THEME.get("alternate_bg", table_bg)
        table_text = CURRENT_THEME.get("table_text", CURRENT_THEME.get("text", "#FFFFFF"))
        active_bg = CURRENT_THEME.get("accent_fill", CURRENT_THEME.get("selection_bg", table_bg))
        header_label_bg = CURRENT_THEME.get("button_hover_bg", alt_bg)

        col_letter = self.column_selector.currentText()
        col_idx = column_index_from_string(col_letter)
        column_indices = []
        headers = []

        if col_idx > 1:
            column_indices.append(col_idx - 1)
            headers.append(get_column_letter(col_idx - 1))

        column_indices.append(col_idx)
        headers.append(col_letter)

        column_indices.append(col_idx + 1)
        headers.append(get_column_letter(col_idx + 1))

        start_row = max(1, self.current_row - 4)
        end_row = min(self.ws.max_row, start_row + 8)
        num_rows = end_row - start_row + 1

        self.preview_table.setRowCount(num_rows)
        self.preview_table.setColumnCount(len(headers))
        self.preview_table.setHorizontalHeaderLabels(headers)

        for r, sheet_row in enumerate(range(start_row, end_row + 1)):
            for c, col in enumerate(column_indices):
                value = self.ws.cell(row=sheet_row, column=col).value
                item = QTableWidgetItem("" if value is None else str(value))
                item.setForeground(QColor(table_text))

                bg_color = alt_bg if r % 2 else table_bg
                if col == col_idx:
                    bg_color = active_bg
                if c == 0:
                    # leftmost column (neighbor) header/title region
                    bg_color = header_label_bg

                item.setBackground(QColor(bg_color))
                self.preview_table.setItem(r, c, item)

        self.preview_table.resizeColumnsToContents()
        self._apply_template_table_theme(self.preview_table)

    def map_event_to_excel(self, row, column):
        if not self.ws or not self.column_selector.currentText():
            return
        try:
            col_letter = self.column_selector.currentText()
            value_raw = self.event_table.item(row, 2).text()
            target_cell = f"{col_letter}{self.current_row}"
            try:
                value = float(value_raw)
            except ValueError:
                value = value_raw
            prev_value = self.ws[target_cell].value
            self.history.append((target_cell, prev_value))
            self.ws[target_cell] = value
            self.wb.save(self.excel_path)
            self.current_row += 1
            self.update_cell_label()
            self.update_preview_table()
        except Exception as e:
            QMessageBox.warning(self, "Mapping Error", f"Failed to assign value: {e}")

    def skip_cell(self):
        self.current_row += 1
        self.update_cell_label()
        self.update_preview_table()

    def undo_last(self):
        if not self.history:
            QMessageBox.information(self, "Undo", "Nothing to undo.")
            return
        cell, old_value = self.history.pop()
        self.ws[cell] = old_value
        self.wb.save(self.excel_path)
        self.current_row = int("".join(filter(str.isdigit, cell)))
        self.update_cell_label()
        self.update_preview_table()

    def finish_and_save(self):
        if self.wb and self.excel_path:
            try:
                self.wb.save(self.excel_path)
                reopen_excel_file_crossplatform(self.excel_path)
                self.accept()
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to save Excel file:\n{e}")


# Auto-update utility


def update_excel_file(excel_path, event_table_data, start_row=3, column_letter="B"):
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        for i, row in enumerate(event_table_data):
            if len(row) < 3:
                continue
            id_val = row[2]
            cell = f"{column_letter}{start_row + i}"
            ws[cell] = id_val
        wb.save(excel_path)
        log.info("Excel file updated with ID values in column %s", column_letter)
    except Exception as e:
        log.error("Failed to update Excel file: %s", e, exc_info=True)


# Cross-platform file reopening logic


def reopen_excel_file_crossplatform(path):
    try:
        time.sleep(1)
        if sys.platform == "darwin":
            safe_path = path.replace("\\", "\\\\").replace('"', '\\"')
            applescript = f"""
            tell application "Microsoft Excel"
                try
                    close (documents whose name is "{os.path.basename(path)}") saving yes
                end try
                open POSIX file "{safe_path}"
                activate
            end tell
            """

            subprocess.call(["osascript", "-e", applescript])
        elif sys.platform == "win32":
            os.startfile(path)
        elif sys.platform.startswith("linux"):
            subprocess.call(["xdg-open", path])
    except Exception as e:
        log.warning("Could not reopen Excel file: %s", e, exc_info=True)
