# VasoAnalyzer
# Copyright © 2025 Osvaldo J. Vega Rodríguez
# Licensed under CC BY-NC-SA 4.0 International
# http://creativecommons.org/licenses/by-nc-sa/4.0/

# coding: utf-8
"""Wizard interface for mapping events into an Excel template.

This module implements a multi-page ``QWizard`` which guides the user
through selecting an Excel template, choosing rows for event mapping and
optionally previewing the result before saving.  The implementation uses
``openpyxl`` to read and write ``.xlsx`` files while preserving formulas
and ``pandas`` for loading event data from CSV files.

The wizard is largely based on the implementation plan documented in the
project repository.
"""

import contextlib
import csv
import hashlib
import json
import os
import re
import tempfile
from collections import Counter, deque
from dataclasses import dataclass, field
from numbers import Real
from pathlib import Path
from typing import Any, cast

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from PyQt6.QtCore import QAbstractTableModel, QMimeData, QModelIndex, QSettings, Qt, pyqtProperty
from PyQt6.QtGui import QBrush, QColor, QFont, QPalette
from PyQt6.QtWidgets import (
    QAbstractItemView,
    QComboBox,
    QFileDialog,
    QHBoxLayout,
    QHeaderView,
    QInputDialog,
    QLabel,
    QListWidget,
    QListWidgetItem,
    QMessageBox,
    QPushButton,
    QTableView,
    QTableWidget,
    QTableWidgetItem,
    QToolButton,
    QVBoxLayout,
    QWizard,
    QWizardPage,
)

import logging

import vasoanalyzer.ui.theme as theme
from vasoanalyzer.excel import TemplateMetadata
from vasoanalyzer.excel.label_matching import normalize_label, best_match

__all__ = ["ExcelMapWizard"]

log = logging.getLogger(__name__)

DEFAULT_QMODEL_INDEX = QModelIndex()

SESSION_EVENT_MIME = "application/vnd.vaso.session-event"



# UI Design Tokens
# ---------------------------------------------------------------------------


def get_semantic_colors() -> dict[str, str]:
    """Get semantic colors from theme."""
    return {
        "success": theme.CURRENT_THEME.get("success_text", "#10B981"),
        "success_bg": theme.CURRENT_THEME.get("success_bg", "#D1FAE5"),
        "warning": theme.CURRENT_THEME.get("warning_text", "#F59E0B"),
        "warning_bg": theme.CURRENT_THEME.get("warning_bg", "#FEF3C7"),
        "error": theme.CURRENT_THEME.get("error_text", "#EF4444"),
        "error_bg": theme.CURRENT_THEME.get("error_bg", "#FEE2E2"),
        "info": theme.CURRENT_THEME.get("info_text", "#3B82F6"),
        "info_bg": theme.CURRENT_THEME.get("info_bg", "#DBEAFE"),
        "muted": theme.CURRENT_THEME.get("muted_text", "#9CA3AF"),
        "muted_bg": theme.CURRENT_THEME.get("muted_bg", "#374151"),
    }


SPACING = {
    "xs": 4,  # Between tightly related items
    "sm": 8,  # Between form elements
    "md": 16,  # Between subsections
    "lg": 24,  # Between major sections
    "xl": 32,  # Between page sections
    "2xl": 48,  # Between distinct areas
}

# Modern UI design tokens
BORDER_RADIUS = {
    "sm": 4,  # Form controls
    "md": 6,  # Buttons
    "lg": 8,  # Cards
}

BUTTON_HEIGHT = 36  # Consistent button height

# Color constants for modern styling
COLORS = {
    "primary": "#0066cc",
    "success": "#28a745",
    "warning": "#ffc107",
    "border_light": "#e0e0e0",
    "border_dark": "#374151",
    "hover_light": "#f0f7ff",
    "hover_dark": "#172554",
}


def get_fonts() -> dict[str, QFont]:
    """Get font styles for different text roles."""
    # Use system font stack for modern typography
    system_font = QFont()
    system_font.setFamily("-apple-system")

    fonts = {
        "h1": QFont(system_font),  # Page titles
        "h2": QFont(system_font),  # Section headers
        "h3": QFont(system_font),  # Subsection headers
        "body": QFont(system_font),  # Normal text
        "small": QFont(system_font),  # Helper text
        "mono": QFont("SF Mono"),  # Code/numbers
    }

    # Configure sizes and weights
    fonts["h1"].setPointSize(16)
    fonts["h1"].setWeight(QFont.Weight.Bold)

    fonts["h2"].setPointSize(14)
    fonts["h2"].setWeight(QFont.Weight.DemiBold)

    fonts["h3"].setPointSize(12)
    fonts["h3"].setWeight(QFont.Weight.DemiBold)

    fonts["body"].setPointSize(14)  # Increased from 11
    fonts["body"].setWeight(QFont.Weight.Normal)

    fonts["small"].setPointSize(12)  # Increased from 10
    fonts["small"].setWeight(QFont.Weight.Normal)

    fonts["mono"].setPointSize(12)  # Increased from 10

    return fonts


def get_modern_table_stylesheet() -> str:
    """Generate modern table stylesheet with theme awareness."""
    base_bg = theme.CURRENT_THEME.get("table_bg", theme.CURRENT_THEME.get("panel_bg", "#FFFFFF"))
    alt_bg = theme.CURRENT_THEME.get("alternate_bg", base_bg)
    border = theme.CURRENT_THEME.get(
        "panel_border",
        theme.CURRENT_THEME.get("table_header_border", "#D1D5DB"),
    )
    hover = theme.CURRENT_THEME.get(
        "table_hover",
        theme.CURRENT_THEME.get("button_hover_bg", alt_bg),
    )
    text = theme.CURRENT_THEME.get("table_text", theme.CURRENT_THEME.get("text", "#111827"))
    header_bg = theme.CURRENT_THEME.get(
        "table_header_bg",
        theme.CURRENT_THEME.get("button_bg", theme.CURRENT_THEME.get("panel_bg", base_bg)),
    )
    selection_bg = theme.CURRENT_THEME.get(
        "selection_bg",
        theme.CURRENT_THEME.get("accent_fill", theme.CURRENT_THEME.get("accent", "#2563EB")),
    )
    selection_text = theme.CURRENT_THEME.get("highlighted_text", "#FFFFFF")

    return f"""
        QTableWidget, QTableView {{
            gridline-color: {border};
            background: {base_bg};
            alternate-background-color: {alt_bg};
            color: {text};
            border: 1px solid {border};
            border-radius: {BORDER_RADIUS["sm"]}px;
            selection-background-color: {selection_bg};
            selection-color: {selection_text};
        }}

        QTableWidget::item, QTableView::item {{
            padding: 8px 12px;
            border: none;
        }}

        QTableWidget::item:hover, QTableView::item:hover {{
            background-color: {hover};
        }}

        QHeaderView::section {{
            background: {header_bg};
            color: {text};
            font-weight: 600;
            font-size: 12px;
            padding: 8px 12px;
            border: none;
            border-right: 1px solid {border};
            border-bottom: 1px solid {border};
        }}

        QHeaderView::section:first {{
            border-top-left-radius: {BORDER_RADIUS["sm"]}px;
        }}

        QHeaderView::section:last {{
            border-top-right-radius: {BORDER_RADIUS["sm"]}px;
            border-right: none;
        }}

        QTableCornerButton::section {{
            background: {header_bg};
            border: 1px solid {border};
        }}
    """


def get_modern_combobox_stylesheet() -> str:
    """Generate modern combobox stylesheet."""
    bg = theme.CURRENT_THEME.get("panel_bg", theme.CURRENT_THEME.get("button_bg", "#FFFFFF"))
    border = theme.CURRENT_THEME.get(
        "panel_border",
        theme.CURRENT_THEME.get("table_header_border", "#D1D5DB"),
    )
    border_focus = theme.CURRENT_THEME.get(
        "accent",
        theme.CURRENT_THEME.get("accent_fill", "#3B82F6"),
    )
    text = theme.CURRENT_THEME.get("text", "#111827")
    hover_bg = theme.CURRENT_THEME.get(
        "table_hover",
        theme.CURRENT_THEME.get("button_hover_bg", bg),
    )
    popup_bg = theme.CURRENT_THEME.get("table_bg", bg)
    selection_bg = theme.CURRENT_THEME.get(
        "selection_bg",
        theme.CURRENT_THEME.get("accent_fill", "#2563EB"),
    )
    selection_text = theme.CURRENT_THEME.get("highlighted_text", "#FFFFFF")

    return f"""
        QComboBox {{
            background-color: {bg};
            border: 1px solid {border};
            border-radius: {BORDER_RADIUS["sm"]}px;
            padding: 6px 8px;
            min-height: 24px;
            color: {text};
            font-size: 14px;
        }}

        QComboBox:hover {{
            background-color: {hover_bg};
        }}

        QComboBox:focus {{
            border: 1px solid {border_focus};
            outline: none;
        }}

        QComboBox::drop-down {{
            border: none;
            width: 20px;
        }}

        QComboBox::down-arrow {{
            width: 12px;
            height: 12px;
        }}

        QComboBox:editable {{
            background-color: {bg};
        }}

        QComboBox QAbstractItemView {{
            background-color: {popup_bg};
            border: 1px solid {border};
            color: {text};
            selection-background-color: {selection_bg};
            selection-color: {selection_text};
        }}
    """


def _get_cell_combobox_stylesheet() -> str:
    """Combobox stylesheet for use inside QTableWidget cells.

    The combo fills the entire cell with no outer border or margin so it
    looks like a native part of the table row.  A chevron on the right
    signals that it is interactive.
    """
    table_bg = theme.CURRENT_THEME.get(
        "table_bg", theme.CURRENT_THEME.get("panel_bg", "#FFFFFF")
    )
    text = theme.CURRENT_THEME.get(
        "table_text", theme.CURRENT_THEME.get("text", "#111827")
    )
    muted = theme.CURRENT_THEME.get("text_disabled", "#9CA3AF")
    hover_bg = theme.CURRENT_THEME.get(
        "table_hover", theme.CURRENT_THEME.get("button_hover_bg", table_bg)
    )
    border_focus = theme.CURRENT_THEME.get(
        "accent", theme.CURRENT_THEME.get("accent_fill", "#3B82F6")
    )
    popup_bg = theme.CURRENT_THEME.get("table_bg", table_bg)
    border = theme.CURRENT_THEME.get(
        "panel_border",
        theme.CURRENT_THEME.get("table_header_border", "#D1D5DB"),
    )
    selection_bg = theme.CURRENT_THEME.get(
        "selection_bg", theme.CURRENT_THEME.get("accent_fill", "#2563EB")
    )
    selection_text = theme.CURRENT_THEME.get("highlighted_text", "#FFFFFF")

    # Inline SVG chevron encoded as a data URI.  The colour is taken from
    # the muted/disabled text token so it stays subtle.
    _c = muted.lstrip("#")
    chevron = (
        f"data:image/svg+xml,"
        f"%3Csvg xmlns='http://www.w3.org/2000/svg' width='10' height='6'%3E"
        f"%3Cpath d='M1 1l4 4 4-4' stroke='%23{_c}' stroke-width='1.5' "
        f"fill='none' stroke-linecap='round' stroke-linejoin='round'/%3E"
        f"%3C/svg%3E"
    )

    return f"""
        QComboBox {{
            background-color: transparent;
            border: none;
            border-radius: 0px;
            padding: 4px 24px 4px 8px;
            margin: 0px;
            color: {text};
            font-size: 14px;
        }}

        QComboBox:hover {{
            background-color: {hover_bg};
        }}

        QComboBox:focus {{
            border: 1px solid {border_focus};
        }}

        QComboBox::drop-down {{
            border: none;
            width: 24px;
            subcontrol-origin: padding;
            subcontrol-position: center right;
        }}

        QComboBox::down-arrow {{
            image: url("{chevron}");
            width: 10px;
            height: 6px;
        }}

        QComboBox QAbstractItemView {{
            background-color: {popup_bg};
            border: 1px solid {border};
            color: {text};
            selection-background-color: {selection_bg};
            selection-color: {selection_text};
        }}
    """


def get_modern_button_stylesheet() -> str:
    """Generate modern button stylesheet with flat design."""
    primary_bg = theme.CURRENT_THEME.get(
        "accent_fill",
        theme.CURRENT_THEME.get("accent", "#2563EB"),
    )
    primary_hover = theme.CURRENT_THEME.get("accent", primary_bg)
    primary_active = theme.CURRENT_THEME.get("button_active_bg", primary_hover)
    secondary_bg = theme.CURRENT_THEME.get(
        "panel_bg",
        theme.CURRENT_THEME.get("button_bg", "#FFFFFF"),
    )
    secondary_hover = theme.CURRENT_THEME.get(
        "table_hover",
        theme.CURRENT_THEME.get("button_hover_bg", secondary_bg),
    )
    secondary_active = theme.CURRENT_THEME.get("button_hover_bg", secondary_hover)
    secondary_border = theme.CURRENT_THEME.get(
        "panel_border",
        theme.CURRENT_THEME.get("table_header_border", "#D1D5DB"),
    )
    text_primary = theme.CURRENT_THEME.get("highlighted_text", "#FFFFFF")
    text_secondary = theme.CURRENT_THEME.get("text", "#111827")
    text_disabled = theme.CURRENT_THEME.get("text_disabled", "#9CA3AF")

    return f"""
        QPushButton {{
            background-color: {secondary_bg};
            color: {text_secondary};
            border: 1px solid {secondary_border};
            border-radius: {BORDER_RADIUS["md"]}px;
            padding: 8px 16px;
            min-height: {BUTTON_HEIGHT}px;
            font-size: 14px;
            font-weight: 500;
        }}

        QPushButton:hover {{
            background-color: {secondary_hover};
        }}

        QPushButton:pressed {{
            background-color: {secondary_active};
        }}

        QPushButton:disabled {{
            color: {text_disabled};
            border-color: {secondary_border};
        }}

        QPushButton#PrimaryButton {{
            background-color: {primary_bg};
            color: {text_primary};
            border: 1px solid {primary_hover};
        }}

        QPushButton#PrimaryButton:hover {{
            background-color: {primary_hover};
        }}

        QPushButton#PrimaryButton:pressed {{
            background-color: {primary_active};
        }}

        QToolButton {{
            background-color: {secondary_bg};
            color: {text_secondary};
            border: 1px solid {secondary_border};
            border-radius: {BORDER_RADIUS["md"]}px;
            padding: 8px;
            min-height: {BUTTON_HEIGHT}px;
        }}

        QToolButton:hover {{
            background-color: {secondary_hover};
        }}

        QToolButton:pressed {{
            background-color: {secondary_active};
        }}

        QToolButton:disabled {{
            color: {text_disabled};
            border-color: {secondary_border};
        }}

        QToolButton#PrimaryButton {{
            background-color: {primary_bg};
            color: {text_primary};
            border: 1px solid {primary_hover};
        }}

        QToolButton#PrimaryButton:hover {{
            background-color: {primary_hover};
        }}

        QToolButton#PrimaryButton:pressed {{
            background-color: {primary_active};
        }}
    """


# Wizard Classes
# ---------------------------------------------------------------------------


class _WizardUnavailableError(RuntimeError):
    """Raised when a wizard page cannot resolve its hosting wizard."""


class WizardPageBase(QWizardPage):
    """QWizardPage helper exposing a typed reference to the host wizard."""

    def _wizard(self) -> "ExcelMapWizard":
        wizard = super().wizard()
        if wizard is None:
            raise _WizardUnavailableError("Wizard is not available")
        return cast("ExcelMapWizard", wizard)


# ---------------------------------------------------------------------------
# Helper utilities
# ---------------------------------------------------------------------------


def load_workbook_preserve(path: str):
    """Load an Excel workbook preserving formulas."""

    keep_vba = Path(path).suffix.lower() == ".xlsm"
    return load_workbook(path, keep_vba=keep_vba, data_only=False)


def load_events_csv(path: str) -> pd.DataFrame:
    """Load a CSV file containing event information."""

    with open(path, encoding="utf-8-sig") as handle:
        sample = handle.read(1024)
        handle.seek(0)
        try:
            delimiter = csv.Sniffer().sniff(sample).delimiter
        except csv.Error:
            if "\t" in sample:
                delimiter = "\t"
            elif ";" in sample:
                delimiter = ";"
            else:
                delimiter = ","

        return pd.read_csv(handle, delimiter=delimiter)


def save_workbook(wb, path: str) -> None:
    """Atomically save an Excel workbook to ``path``."""

    target_path = Path(path)
    temp_fd, temp_name = tempfile.mkstemp(
        prefix=f".{target_path.stem}.",
        suffix=target_path.suffix,
        dir=str(target_path.parent),
    )
    os.close(temp_fd)

    temp_path = Path(temp_name)
    try:
        wb.save(str(temp_path))
        os.replace(str(temp_path), str(target_path))
    finally:
        if temp_path.exists():
            with contextlib.suppress(OSError):
                temp_path.unlink()


# ---------------------------------------------------------------------------
# Model for previewing pandas.DataFrame in a QTableView
# ---------------------------------------------------------------------------


class PandasModel(QAbstractTableModel):
    """Simple table model exposing a pandas DataFrame."""

    def __init__(self, frame: pd.DataFrame):
        super().__init__()
        self._df = frame.reset_index(drop=True)

    def rowCount(self, parent: QModelIndex = DEFAULT_QMODEL_INDEX) -> int:
        return len(self._df)

    def columnCount(self, parent: QModelIndex = DEFAULT_QMODEL_INDEX) -> int:
        return len(self._df.columns)

    def data(self, index: QModelIndex, role: int = Qt.ItemDataRole.DisplayRole):
        if not index.isValid() or role != Qt.ItemDataRole.DisplayRole:
            return None
        value = self._df.iat[index.row(), index.column()]
        return "" if pd.isna(value) else str(value)

    def headerData(self, section: int, orientation: Qt.Orientation, role: int = Qt.ItemDataRole.DisplayRole):
        if role != Qt.ItemDataRole.DisplayRole:
            return None
        if orientation == Qt.Orientation.Horizontal:
            return self._df.columns[section]
        return str(section + 1)


class SessionValuesTable(QTableWidget):
    """Table showing session events and acting as a drag source."""

    def __init__(self, parent=None) -> None:
        super().__init__(parent)
        self.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.setAlternatingRowColors(True)
        self.setDragEnabled(True)
        self.setDragDropMode(QAbstractItemView.DragDropMode.DragOnly)
        self.verticalHeader().setVisible(False)

    def mimeTypes(self) -> list[str]:
        return [SESSION_EVENT_MIME, "text/plain"]

    def mimeData(self, indexes):  # noqa: N802 - Qt API
        mime = QMimeData()
        if not indexes:
            return mime

        row = indexes[0].row()
        item = self.item(row, 0)
        if item is None:
            return mime

        event_index = item.data(Qt.ItemDataRole.UserRole)
        if event_index is None:
            return mime

        mime.setData(SESSION_EVENT_MIME, str(int(event_index)).encode())
        mime.setText(item.text())
        return mime


class TemplatePreviewTable(QTableWidget):
    """Preview table that accepts drops on the active measurement column."""

    def __init__(self, parent=None) -> None:
        super().__init__(parent)
        self._drop_context = None
        self._active_drop_column: int | None = None
        self.setAcceptDrops(True)
        self.setDragDropMode(QAbstractItemView.DragDropMode.DropOnly)
        self.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self._apply_theme()

    def set_drop_context(self, context) -> None:
        self._drop_context = context

    def set_active_drop_column(self, column_index: int | None) -> None:
        self._active_drop_column = column_index

    def _extract_drop_target(self, event) -> tuple[int, int] | None:
        if self._drop_context is None:
            return None

        event_index = self._drop_context.event_index_from_mime(event.mimeData())
        if event_index is None:
            return None

        index = self.indexAt(event.pos())
        if not index.isValid():
            return None

        if self._active_drop_column is None or index.column() != self._active_drop_column:
            return None

        template_row = self._drop_context.template_row_for_preview_row(index.row())
        if template_row is None:
            return None

        return template_row, event_index

    def dragEnterEvent(self, event):  # noqa: N802 - Qt API
        if self._extract_drop_target(event):
            event.acceptProposedAction()
        else:
            event.ignore()

    def dragMoveEvent(self, event):  # noqa: N802 - Qt API
        if self._extract_drop_target(event):
            event.acceptProposedAction()
        else:
            event.ignore()

    def dropEvent(self, event):  # noqa: N802 - Qt API
        target = self._extract_drop_target(event)
        if not target:
            event.ignore()
            return

        template_row, event_index = target
        self._drop_context.assign_event_to_row(template_row, event_index)
        event.acceptProposedAction()

    def _apply_theme(self) -> None:
        """Apply theme-driven palette for dark/light modes."""
        palette = self.palette()
        table_bg = theme.CURRENT_THEME.get("table_bg", "#020617")
        alt_bg = theme.CURRENT_THEME.get("alternate_bg", table_bg)
        text = theme.CURRENT_THEME.get("table_text", theme.CURRENT_THEME.get("text", "#FFFFFF"))
        highlight = theme.CURRENT_THEME.get("selection_bg", "#1D4ED8")
        highlighted_text = theme.CURRENT_THEME.get("highlighted_text", "#FFFFFF")

        palette.setColor(self.backgroundRole(), QColor(table_bg))
        palette.setColor(self.foregroundRole(), QColor(text))
        palette.setColor(QPalette.ColorRole.Base, QColor(table_bg))
        palette.setColor(QPalette.ColorRole.AlternateBase, QColor(alt_bg))
        palette.setColor(QPalette.ColorRole.Text, QColor(text))
        palette.setColor(QPalette.ColorRole.WindowText, QColor(text))
        palette.setColor(QPalette.ColorRole.Highlight, QColor(highlight))
        palette.setColor(QPalette.ColorRole.HighlightedText, QColor(highlighted_text))

        self.setPalette(palette)
        # Use the new modern table stylesheet
        self.setStyleSheet(get_modern_table_stylesheet())

    def apply_theme(self) -> None:
        """Public hook to refresh palette after theme changes."""

        self._apply_theme()


# ---------------------------------------------------------------------------
# Wizard Pages
# ---------------------------------------------------------------------------


class TemplatePage(WizardPageBase):
    """Page for selecting the Excel template and event CSV."""

    def __init__(self) -> None:
        super().__init__()
        self.setTitle("Step 1: Load Template & CSV")
        layout = QVBoxLayout(self)
        layout.setSpacing(SPACING["md"])
        layout.setContentsMargins(SPACING["lg"], SPACING["lg"], SPACING["lg"], SPACING["lg"])

        self._templatePath = ""
        self._csvPath = ""
        self.registerField("templatePath*", self, "templatePath")
        self.registerField("csvPath*", self, "csvPath")

        self.btn_excel = QPushButton("Load Excel Template…")
        self.btn_excel.setMinimumHeight(BUTTON_HEIGHT)
        self.btn_excel.setStyleSheet(get_modern_button_stylesheet())
        self.lbl_excel = QLabel("No template loaded.")
        fonts = get_fonts()
        self.lbl_excel.setFont(fonts["body"])
        self.btn_excel.clicked.connect(self.load_template)
        layout.addWidget(self.btn_excel)
        layout.addWidget(self.lbl_excel)

        # Sheet selector UI
        self.lbl_sheet = QLabel("Select worksheet:")
        self.combo_sheet = QComboBox()
        self.combo_sheet.setStyleSheet(get_modern_combobox_stylesheet())
        self.lbl_sheet.setVisible(False)
        self.combo_sheet.setVisible(False)
        self.combo_sheet.currentIndexChanged.connect(self._on_sheet_changed)
        layout.addWidget(self.lbl_sheet)
        layout.addWidget(self.combo_sheet)

        # Recent templates UI
        self.recent_templates_label = QLabel("Recent templates")
        self.recent_templates_list = QListWidget()
        self.recent_templates_list.setSelectionMode(QListWidget.SelectionMode.SingleSelection)
        self.recent_templates_list.setAlternatingRowColors(True)
        self.recent_templates_list.setUniformItemSizes(True)
        self.recent_templates_list.itemActivated.connect(self._on_recent_template_activated)
        self._apply_recent_templates_style()
        self.remove_recent_button = QPushButton("Remove selected")
        self.remove_recent_button.setMinimumHeight(BUTTON_HEIGHT)
        self.remove_recent_button.setStyleSheet(get_modern_button_stylesheet())
        self.clear_recent_button = QPushButton("Clear all")
        self.clear_recent_button.setMinimumHeight(BUTTON_HEIGHT)
        self.clear_recent_button.setStyleSheet(get_modern_button_stylesheet())
        self.remove_recent_button.clicked.connect(self._on_remove_selected_recent_template)
        self.clear_recent_button.clicked.connect(self._on_clear_recent_templates)
        self.recent_templates_label.setVisible(False)
        self.recent_templates_list.setVisible(False)
        self.remove_recent_button.setVisible(False)
        self.clear_recent_button.setVisible(False)
        layout.addWidget(self.recent_templates_label)
        layout.addWidget(self.recent_templates_list)
        recent_buttons_row = QHBoxLayout()
        recent_buttons_row.setSpacing(SPACING["sm"])
        recent_buttons_row.addStretch()
        recent_buttons_row.addWidget(self.remove_recent_button)
        recent_buttons_row.addWidget(self.clear_recent_button)
        layout.addLayout(recent_buttons_row)

        # Add extra spacing before CSV section
        layout.addSpacing(SPACING["xl"])

        self.btn_csv = QPushButton("Load Events CSV…")
        self.btn_csv.setMinimumHeight(BUTTON_HEIGHT)
        self.btn_csv.setStyleSheet(get_modern_button_stylesheet())
        self.lbl_csv = QLabel("No events loaded.")
        self.lbl_csv.setFont(fonts["body"])
        self.btn_csv.clicked.connect(self.load_csv)
        layout.addWidget(self.btn_csv)
        layout.addWidget(self.lbl_csv)

        self._update_recent_templates_list()

    def _apply_recent_templates_style(self) -> None:
        list_bg = theme.CURRENT_THEME.get(
            "table_bg",
            theme.CURRENT_THEME.get("panel_bg", "#FFFFFF"),
        )
        list_border = theme.CURRENT_THEME.get(
            "panel_border",
            theme.CURRENT_THEME.get("table_header_border", "#D1D5DB"),
        )
        list_hover = theme.CURRENT_THEME.get(
            "table_hover",
            theme.CURRENT_THEME.get("button_hover_bg", list_bg),
        )
        list_text = theme.CURRENT_THEME.get(
            "table_text",
            theme.CURRENT_THEME.get("text", "#111827"),
        )
        selection_bg = theme.CURRENT_THEME.get(
            "selection_bg",
            theme.CURRENT_THEME.get("accent_fill", "#2563EB"),
        )
        selection_text = theme.CURRENT_THEME.get("highlighted_text", "#FFFFFF")
        self.recent_templates_list.setStyleSheet(f"""
            QListWidget {{
                border: 1px solid {list_border};
                border-radius: {BORDER_RADIUS["sm"]}px;
                background-color: {list_bg};
                padding: 4px;
                color: {list_text};
            }}
            QListWidget::item {{
                padding: 8px 12px;
                border-radius: {BORDER_RADIUS["sm"]}px;
            }}
            QListWidget::item:hover {{
                background-color: {list_hover};
            }}
            QListWidget::item:selected {{
                background-color: {selection_bg};
                color: {selection_text};
            }}
        """)

    def apply_theme(self) -> None:
        self.btn_excel.setStyleSheet(get_modern_button_stylesheet())
        self.btn_csv.setStyleSheet(get_modern_button_stylesheet())
        self.combo_sheet.setStyleSheet(get_modern_combobox_stylesheet())
        self.remove_recent_button.setStyleSheet(get_modern_button_stylesheet())
        self.clear_recent_button.setStyleSheet(get_modern_button_stylesheet())
        self._apply_recent_templates_style()

    # Properties exposed as wizard fields
    def get_templatePath(self) -> str:
        return self._templatePath

    def set_templatePath(self, value: str) -> None:
        self._templatePath = value

    templatePath = pyqtProperty(str, fget=get_templatePath, fset=set_templatePath)

    def get_csvPath(self) -> str:
        return self._csvPath

    def set_csvPath(self, value: str) -> None:
        self._csvPath = value

    csvPath = pyqtProperty(str, fget=get_csvPath, fset=set_csvPath)

    # ------------------------------------------------------
    def initializePage(self) -> None:
        super().initializePage()
        self._update_events_status()

    # ------------------------------------------------------
    def _load_template_from_path(self, path: str) -> None:
        """Load a template from the given path and update wizard state."""
        if not path:
            return

        log.info(f"[Wizard] Loading template: {path}")
        try:
            wb = load_workbook_preserve(path)
        except Exception as exc:  # pragma: no cover - GUI feedback
            log.error(f"[Wizard] Failed to open workbook: {exc}", exc_info=True)
            QMessageBox.critical(self, "Load Failed", str(exc))
            return

        log.info(f"[Wizard] Workbook opened. Sheets: {wb.sheetnames}")

        # Don't set ws yet - wait for sheet selection
        wiz = self._wizard()
        wiz.setField("templatePath", path)
        wiz.wb = wb
        wiz.ws = None  # Will be set after sheet selection
        wiz.selected_sheet_name = None
        wiz.reset_mapping_state()

        # Populate sheet selector
        self._populate_sheet_selector(wb)

        # Try to restore previous sheet preference
        saved_sheet = self._load_sheet_preference(path)
        if saved_sheet and saved_sheet in wb.sheetnames:
            idx = self.combo_sheet.findText(saved_sheet)
            if idx >= 0:
                self.combo_sheet.setCurrentIndex(idx)
        elif self.combo_sheet.count() == 1:
            # Auto-select if only one sheet
            self.combo_sheet.setCurrentIndex(0)
        elif self.combo_sheet.count() > 1:
            # Show selection required message
            self.lbl_excel.setText(
                f"Loaded: {Path(path).name} - Please select a worksheet to continue"
            )
            colors = get_semantic_colors()
            self.lbl_excel.setStyleSheet(f"color: {colors['warning']};")

            # Set placeholder text on combo box for clarity
            self.combo_sheet.setPlaceholderText("Select a worksheet...")

            # Ensure currentIndex is -1 (nothing selected)
            if self.combo_sheet.currentIndex() == 0:
                self.combo_sheet.setCurrentIndex(-1)
        else:
            # No sheets available (shouldn't happen)
            self.lbl_excel.setText(f"Loaded: {Path(path).name} - No worksheets found")
            colors = get_semantic_colors()
            self.lbl_excel.setStyleSheet(f"color: {colors['error']};")

        # Always call _on_sheet_changed explicitly for the selected sheet.
        # setCurrentIndex() above does NOT guarantee currentIndexChanged fires
        # when the combo was pre-populated with signals blocked (blockSignals=True
        # in _populate_sheet_selector), leaving wiz.ws=None and Next grayed out.
        current_idx = self.combo_sheet.currentIndex()
        if current_idx >= 0:
            self._on_sheet_changed(current_idx)

        self._update_recent_templates(path)
        self.completeChanged.emit()

    # ------------------------------------------------------
    def load_template(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Excel Template",
            "",
            "Excel Files (*.xlsx *.xlsm);;Macro-Enabled (*.xlsm);;Standard (*.xlsx)",
        )
        if not path:
            return
        self._load_template_from_path(path)

    # ------------------------------------------------------
    def load_csv(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "Select CSV", "", "CSV Files (*.csv)")
        if not path:
            return
        try:
            df = self._prepare_events_dataframe(load_events_csv(path))
        except ValueError as exc:
            QMessageBox.warning(self, "Invalid CSV", str(exc))
            return
        except Exception as exc:  # pragma: no cover - GUI feedback
            QMessageBox.critical(self, "Load Failed", str(exc))
            return

        wiz = self._wizard()
        wiz.reset_mapping_state()
        wiz.setField("csvPath", path)
        wiz.set_events_dataframe(df, source="csv")
        self._update_events_status()
        self.completeChanged.emit()

    # ------------------------------------------------------
    def isComplete(self) -> bool:
        wiz = self._wizard()
        has_events = bool(self.field("csvPath")) or getattr(wiz, "eventsDF", None) is not None
        has_sheet_selected = wiz.ws is not None  # Check sheet selection
        return bool(self.field("templatePath") and has_events and has_sheet_selected)

    # ------------------------------------------------------
    def _update_events_status(self) -> None:
        wiz = self._wizard()
        df = getattr(wiz, "eventsDF", None)
        if df is None or df.empty:
            self.lbl_csv.setText("No events loaded.")
            self.btn_csv.setText("Load Events CSV…")
            return

        source = getattr(wiz, "events_source", "session")
        count = len(df)
        noun = "event" if count == 1 else "events"
        if source == "session":
            message = f"Using {count} {noun} from current session."
            self.btn_csv.setText("Replace Events CSV…")
        else:
            csv_path = Path(self.field("csvPath") or "")
            base = csv_path.name if csv_path.name else "external CSV"
            message = f"Loaded {count} {noun} from {base}."
            self.btn_csv.setText("Reload Events CSV…")
        self.lbl_csv.setText(message)

    # ------------------------------------------------------
    def _load_recent_templates(self) -> list[str]:
        """Return the stored recent templates list."""
        settings = self._get_settings()
        paths = settings.value("recentTemplates", [], type=list) or []
        return [p for p in paths if p]

    # ------------------------------------------------------
    def _save_recent_templates(self, paths: list[str]) -> None:
        """Persist the recent templates list."""
        settings = self._get_settings()
        settings.setValue("recentTemplates", paths)

    # ------------------------------------------------------
    def _update_recent_templates_list(self) -> None:
        """Refresh the Recent Templates UI from QSettings."""
        self.recent_templates_list.clear()
        paths = self._load_recent_templates()
        has_items = bool(paths)

        if not has_items:
            self.recent_templates_label.setVisible(False)
            self.recent_templates_list.setVisible(False)
            self.remove_recent_button.setVisible(False)
            self.clear_recent_button.setVisible(False)
            return

        for path in paths:
            file_name = os.path.basename(path)
            item = QListWidgetItem(file_name)
            item.setToolTip(path)
            item.setData(Qt.ItemDataRole.UserRole, path)
            self.recent_templates_list.addItem(item)

        self.recent_templates_label.setVisible(True)
        self.recent_templates_list.setVisible(True)
        self.remove_recent_button.setVisible(True)
        self.clear_recent_button.setVisible(True)

    # ------------------------------------------------------
    def _update_recent_templates(self, path: str) -> None:
        """Update the stored recent templates list and refresh the UI."""
        if not path:
            return

        paths = self._load_recent_templates()
        if path in paths:
            paths.remove(path)
        paths.insert(0, path)
        paths = paths[:5]
        self._save_recent_templates(paths)
        self._update_recent_templates_list()

    # ------------------------------------------------------
    def _on_recent_template_activated(self, item: QListWidgetItem) -> None:
        """Handle activation of a recent template entry."""
        if item is None:
            return
        path = item.data(Qt.ItemDataRole.UserRole)
        if not path:
            return
        self._load_template_from_path(path)

    # ------------------------------------------------------
    def _on_clear_recent_templates(self) -> None:
        """Clear all recent templates."""
        self._save_recent_templates([])
        self._update_recent_templates_list()

    # ------------------------------------------------------
    def _on_remove_selected_recent_template(self) -> None:
        """Remove only the currently selected recent template."""
        item = self.recent_templates_list.currentItem()
        if item is None:
            return
        path = item.data(Qt.ItemDataRole.UserRole)
        if not path:
            return
        paths = [p for p in self._load_recent_templates() if p != path]
        self._save_recent_templates(paths)
        self._update_recent_templates_list()

    # ------------------------------------------------------
    @staticmethod
    def _get_settings() -> QSettings:
        return QSettings("TykockiLab", "VasoAnalyzer")

    # ------------------------------------------------------
    def _load_sheet_preference(self, file_path: str) -> str | None:
        """Load previously selected sheet for this file."""
        settings = self._get_settings()
        key = f"excel_sheet_selection/{file_path}"
        return settings.value(key, None, type=str)

    # ------------------------------------------------------
    def _save_sheet_preference(self, file_path: str, sheet_name: str) -> None:
        """Save sheet selection preference."""
        settings = self._get_settings()
        key = f"excel_sheet_selection/{file_path}"
        settings.setValue(key, sheet_name)

    # ------------------------------------------------------
    def _populate_sheet_selector(self, wb) -> None:
        """Populate sheet selector with workbook sheets."""

        self.combo_sheet.blockSignals(True)
        self.combo_sheet.clear()

        for sheet_name in wb.sheetnames:
            # Skip hidden metadata sheets
            if sheet_name.startswith("VasoMetadata"):
                continue
            # Skip hidden sheets
            sheet = wb[sheet_name]
            if hasattr(sheet, "sheet_state") and sheet.sheet_state == "hidden":
                continue
            self.combo_sheet.addItem(sheet_name)

        self.combo_sheet.blockSignals(False)

        # Clear auto-selection - force user to explicitly choose when multiple sheets
        if self.combo_sheet.count() > 1:
            self.combo_sheet.setCurrentIndex(-1)

        # Show selector if sheets available
        has_sheets = self.combo_sheet.count() > 0
        self.lbl_sheet.setVisible(has_sheets)
        self.combo_sheet.setVisible(has_sheets)

    # ------------------------------------------------------
    def _on_sheet_changed(self, index: int) -> None:
        """Handle sheet selection change."""
        if index < 0:
            return

        sheet_name = self.combo_sheet.currentText()
        if not sheet_name:
            return

        wiz = self._wizard()
        template_path = self.field("templatePath")

        if not wiz.wb:
            return

        # Update wizard state
        wiz.ws = wiz.wb[sheet_name]
        wiz.selected_sheet_name = sheet_name
        wiz.reset_mapping_state()

        # Save preference
        if template_path:
            self._save_sheet_preference(template_path, sheet_name)

        # Reload metadata for new sheet
        self._reload_metadata_for_sheet(sheet_name)

        # Update UI label
        from pathlib import Path

        self.lbl_excel.setText(f"Loaded: {Path(template_path).name} (Sheet: {sheet_name})")
        self.lbl_excel.setStyleSheet("")

        self.completeChanged.emit()

    # ------------------------------------------------------
    def _reload_metadata_for_sheet(self, sheet_name: str) -> None:
        """Reload metadata for the selected sheet.

        Priority:
        1. Sheet-specific VasoMetadata_{sheet} hidden sheet
        2. Full inference via read_template_metadata (named ranges → structure)
        """
        from vasoanalyzer.excel.template_metadata import (
            read_sheet_specific_metadata,
            read_template_metadata,
        )

        wiz = self._wizard()
        if not wiz.wb:
            log.warning("[Wizard] _reload_metadata_for_sheet called but wiz.wb is None")
            return

        log.info(f"[Wizard] Loading metadata for sheet '{sheet_name}'")
        try:
            # 1. Try sheet-specific VasoMetadata first
            metadata = read_sheet_specific_metadata(wiz.wb, sheet_name)
            if metadata:
                log.info(f"[Wizard] Found sheet-specific VasoMetadata for '{sheet_name}'")
            else:
                log.debug(f"[Wizard] No VasoMetadata sheet for '{sheet_name}', trying inference")

            # 2. Fall back to full inference (named ranges → auto-detect structure)
            if not metadata:
                template_path = self.field("templatePath")
                metadata = read_template_metadata(template_path, wb=wiz.wb, sheet_name=sheet_name)

            if metadata:
                log.info(
                    f"[Wizard] Metadata ready: source={metadata.source!r}, "
                    f"date_row={metadata.date_row}, "
                    f"event_rows={len(metadata.event_rows)}, "
                    f"date_cols={len(metadata.date_columns)}"
                )
                wiz.template_metadata = metadata
                template_path = self.field("templatePath")
                _SOURCE_LABELS = {
                    "sheet_specific": "VasoMetadata sheet",
                    "vba_metadata": "VasoMetadata sheet",
                    "named_ranges": "named ranges",
                    "inferred": "auto-detected",
                }
                source_label = _SOURCE_LABELS.get(metadata.source, metadata.source)
                status_msg = (
                    f"✓ Loaded: {Path(template_path).name}"
                    f" (Sheet: {sheet_name}, {source_label})"
                )
                self.lbl_excel.setText(status_msg)
                colors = get_semantic_colors()
                self.lbl_excel.setStyleSheet(f"color: {colors['success']};")
            else:
                log.warning(
                    f"[Wizard] No metadata detected for sheet '{sheet_name}'. "
                    "prepare_layout will try named ranges as final fallback."
                )
        except Exception as exc:
            log.error(f"[Wizard] Metadata load error for '{sheet_name}': {exc}", exc_info=True)
            template_path = self.field("templatePath")
            QMessageBox.warning(
                self, "Metadata Error", f"Could not load metadata for sheet '{sheet_name}':\n{exc}"
            )
            self.lbl_excel.setText(
                f"Loaded: {Path(template_path).name} (Sheet: {sheet_name}, metadata error)"
            )
            colors = get_semantic_colors()
            self.lbl_excel.setStyleSheet(f"color: {colors['warning']};")

    # ------------------------------------------------------
    @staticmethod
    def _prepare_events_dataframe(df: pd.DataFrame) -> pd.DataFrame:
        if df is None or df.empty:
            raise ValueError("Event CSV is empty.")

        rename_map: dict[str, str] = {}

        def _norm(col: str) -> str:
            return "".join(ch for ch in col.lower() if ch.isalnum())

        for col in df.columns:
            norm = _norm(col)
            if norm in {"event", "eventlabel", "label"}:
                rename_map[col] = "Event"
            elif norm in {"time", "times", "seconds", "timeseconds"}:
                rename_map[col] = "Time (s)"
            elif norm in {"id", "innerdiameter", "idum", "idmicrometer", "diameter"}:
                rename_map[col] = "ID (µm)"
            elif norm in {"od", "outerdiameter", "odum", "odmicrometer"}:
                rename_map[col] = "OD (µm)"
            elif norm in {"frame", "framenumber", "frameindex"}:
                rename_map[col] = "Frame"

        if rename_map:
            df = df.rename(columns=rename_map)

        if "Event" not in df.columns:
            raise ValueError("Event CSV must contain an 'Event' column.")

        value_candidates = [col for col in df.columns if col in {"ID (µm)", "OD (µm)"}]
        if not value_candidates:
            raise ValueError("Event CSV must include 'ID (µm)' and/or 'OD (µm)' column.")

        desired_order = [
            col for col in ["Event", "Time (s)", "ID (µm)", "OD (µm)", "Frame"] if col in df.columns
        ]
        df = df[desired_order].copy()
        df["Event"] = df["Event"].astype(str)

        for col in ["Time (s)", "ID (µm)", "OD (µm)"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")
        if "Frame" in df.columns:
            df["Frame"] = pd.to_numeric(df["Frame"], errors="coerce")

        return df.reset_index(drop=True)


@dataclass
class SessionEventInfo:
    index: int
    label: str
    time_value: float | None
    values: dict[str, Any] = field(default_factory=dict)

    @property
    def combo_text(self) -> str:
        if self.time_value is None or pd.isna(self.time_value):
            return self.label
        return f"{self.label} ({self.time_value:.2f}s)"


@dataclass
class EventRowInfo:
    row_index: int
    label: str
    is_header: bool
    label_cell: str


@dataclass
class DateColumnOption:
    column_index: int
    cell_address: str
    value: Any
    empty_slots: int
    is_new: bool = False

    @property
    def letter(self) -> str:
        return str(get_column_letter(self.column_index))

    @property
    def display(self) -> str:
        base = self.letter
        if self.value not in (None, ""):
            from datetime import datetime

            val = self.value
            if isinstance(val, datetime):
                val = val.strftime("%m/%d/%Y")
            base = f"{base} – {val}"
        if self.is_new:
            base += " (new)"
        elif self.empty_slots == 0:
            base += " (full)"
        return base


class RowMappingPage(WizardPageBase):
    """Interactive mapping page with preview and row-by-row controls."""

    def __init__(self) -> None:
        super().__init__()
        self.setTitle("Step 2: Map Events to Template Rows")

        self._event_row_widgets: dict[int, QComboBox] = {}
        self._value_items: dict[int, QTableWidgetItem] = {}
        self._current_value_items: dict[int, QTableWidgetItem] = {}
        self._status_items: dict[int, QTableWidgetItem] = {}
        self._template_row_to_table_index: dict[int, int] = {}
        self._initialised = False

        root = QVBoxLayout(self)
        root.setSpacing(SPACING["sm"])
        root.setContentsMargins(SPACING["lg"], SPACING["md"], SPACING["lg"], SPACING["md"])

        fonts = get_fonts()
        colors = get_semantic_colors()
        border_color = theme.CURRENT_THEME.get(
            "panel_border",
            theme.CURRENT_THEME.get("table_header_border", "#D1D5DB"),
        )
        muted_text = colors["muted"]

        # ---- status banner ----
        self.info_label = QLabel()
        self.info_label.setWordWrap(True)
        self.info_label.setFont(fonts["body"])
        root.addWidget(self.info_label)

        # ---- control bar ----
        control_row = QHBoxLayout()
        control_row.setSpacing(SPACING["sm"])
        control_row.setContentsMargins(0, 0, 0, 0)

        meas_label = QLabel("Measurement:")
        meas_label.setFont(fonts["small"])
        control_row.addWidget(meas_label)
        self.measurement_combo = QComboBox()
        self.measurement_combo.setStyleSheet(get_modern_combobox_stylesheet())
        control_row.addWidget(self.measurement_combo)

        self.pick_date_combo = QComboBox()
        self.pick_date_combo.setStyleSheet(get_modern_combobox_stylesheet())
        self.pick_date_combo.setVisible(False)
        control_row.addWidget(self.pick_date_combo)

        self.redetect_btn = QToolButton()
        self.redetect_btn.setText("Re-detect")
        self.redetect_btn.setStyleSheet(get_modern_button_stylesheet())
        control_row.addWidget(self.redetect_btn)

        self.select_unmapped_btn = QToolButton()
        self.select_unmapped_btn.setText("Select Unmapped…")
        self.select_unmapped_btn.setObjectName("PrimaryButton")
        self.select_unmapped_btn.setStyleSheet(get_modern_button_stylesheet())
        self.select_unmapped_btn.setVisible(False)
        control_row.addWidget(self.select_unmapped_btn)

        control_row.addStretch()
        root.addLayout(control_row)

        # ---- mapping table (full width) ----
        self.mapping_table = QTableWidget()
        self.mapping_table.setColumnCount(6)
        self.mapping_table.setHorizontalHeaderLabels(
            ["Row", "Template Label", "Session Event", "Value", "Current", ""]
        )
        self.mapping_table.verticalHeader().setVisible(False)
        self.mapping_table.verticalHeader().setDefaultSectionSize(40)
        self.mapping_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._apply_table_theme(self.mapping_table)
        self.mapping_table.setAlternatingRowColors(False)
        mh = self.mapping_table.horizontalHeader()
        mh.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        mh.resizeSection(0, 48)
        mh.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        mh.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        mh.setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)
        mh.resizeSection(3, 80)
        mh.setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)
        mh.resizeSection(4, 80)
        mh.setSectionResizeMode(5, QHeaderView.ResizeMode.Fixed)
        mh.resizeSection(5, 40)
        root.addWidget(self.mapping_table, 1)

        # ---- helper text ----
        helper_lines = [
            "Headers use bold, filled Column A cells; the wizard never writes to them.",
            "To add an event row, type its label in Column A with normal text and no fill.",
            "Pick the active date column in the row of dates; values go into that column.",
            "Override matches using the dropdowns — your selections control the final export.",
        ]
        helper_text = QLabel("  •  ".join(helper_lines))
        helper_text.setWordWrap(True)
        helper_text.setFont(fonts["small"])
        helper_text.setStyleSheet(
            f"color: {muted_text};"
            f"border-top: 1px solid {border_color};"
            f"padding-top: {SPACING['sm']}px;"
        )
        root.addWidget(helper_text)

        self.measurement_combo.currentTextChanged.connect(self._on_measurement_changed)
        self.redetect_btn.clicked.connect(self._on_redetect)
        self.pick_date_combo.currentIndexChanged.connect(self._on_date_changed)
        self.select_unmapped_btn.clicked.connect(self._select_all_unmapped)

    def _apply_table_theme(self, table) -> None:
        """Apply theme-aware palette and stylesheet to table widgets."""
        palette = table.palette()
        table_bg = theme.CURRENT_THEME.get("table_bg", "#020617")
        alt_bg = theme.CURRENT_THEME.get("alternate_bg", table_bg)
        text = theme.CURRENT_THEME.get("table_text", theme.CURRENT_THEME.get("text", "#FFFFFF"))
        highlight = theme.CURRENT_THEME.get("selection_bg", "#1D4ED8")
        highlighted_text = theme.CURRENT_THEME.get("highlighted_text", "#FFFFFF")

        palette.setColor(QPalette.ColorRole.Base, QColor(table_bg))
        palette.setColor(QPalette.ColorRole.AlternateBase, QColor(alt_bg))
        palette.setColor(QPalette.ColorRole.Text, QColor(text))
        palette.setColor(QPalette.ColorRole.WindowText, QColor(text))
        palette.setColor(QPalette.ColorRole.Highlight, QColor(highlight))
        palette.setColor(QPalette.ColorRole.HighlightedText, QColor(highlighted_text))

        table.setPalette(palette)
        # Use the new modern table stylesheet
        table.setStyleSheet(get_modern_table_stylesheet())

    def apply_theme(self) -> None:
        self.measurement_combo.setStyleSheet(get_modern_combobox_stylesheet())
        self.pick_date_combo.setStyleSheet(get_modern_combobox_stylesheet())
        self.redetect_btn.setStyleSheet(get_modern_button_stylesheet())
        self.select_unmapped_btn.setStyleSheet(get_modern_button_stylesheet())
        self._apply_table_theme(self.mapping_table)

    # --------------------------------------------------
    def initializePage(self) -> None:
        super().initializePage()
        wiz = self._wizard()
        if not wiz or wiz.wb is None or wiz.ws is None or wiz.eventsDF is None:
            self.info_label.setText("Load an Excel template and event data first.")
            self.setFinalPage(True)
            return

        if not wiz.prepare_layout(auto=True):
            self.info_label.setText(getattr(wiz, "layout_error", ""))
            return

        self._populate_measurement_options()
        self._populate_date_options()
        self._rebuild_mapping_table()
        self._update_status_banner()
        self._initialised = True

    # --------------------------------------------------
    def _populate_measurement_options(self) -> None:
        wiz = self._wizard()
        self.measurement_combo.blockSignals(True)
        self.measurement_combo.clear()
        for col in wiz.measurement_columns:
            self.measurement_combo.addItem(col)
        if self.measurement_combo.count():
            if wiz.current_measurement:
                idx = self.measurement_combo.findText(wiz.current_measurement)
                if idx >= 0:
                    self.measurement_combo.setCurrentIndex(idx)
            wiz.current_measurement = self.measurement_combo.currentText()
        self.measurement_combo.blockSignals(False)

    # --------------------------------------------------
    def _populate_date_options(self) -> None:
        wiz = self._wizard()
        options = wiz.date_columns or []

        self.pick_date_combo.blockSignals(True)
        self.pick_date_combo.clear()
        for opt in options:
            self.pick_date_combo.addItem(opt.display, opt)

        if not options:
            self.pick_date_combo.setVisible(False)
            return

        current = wiz.active_date_column
        if current:
            for idx, opt in enumerate(options):
                if opt.column_index == current.column_index:
                    self.pick_date_combo.setCurrentIndex(idx)
                    break

        needs_choice = wiz.manual_date_selection_required or len(options) > 1
        self.pick_date_combo.setVisible(needs_choice)
        self.pick_date_combo.blockSignals(False)
        wiz.ensure_active_date_value(self)

    # --------------------------------------------------
    def _rebuild_mapping_table(self) -> None:
        wiz = self._wizard()
        event_rows = [row for row in wiz.event_rows if not row.is_header]

        self.mapping_table.setRowCount(len(event_rows))
        self._event_row_widgets.clear()
        self._value_items.clear()
        self._current_value_items: dict[int, QTableWidgetItem] = {}
        self._status_items.clear()
        self._template_row_to_table_index = {}

        font_mono = QFont("Menlo", 10)
        no_edit = Qt.ItemFlag.ItemIsEditable

        for row_idx, event_row in enumerate(event_rows):
            row_number_item = QTableWidgetItem(str(event_row.row_index))
            row_number_item.setFlags(row_number_item.flags() & ~no_edit)
            row_number_item.setFont(font_mono)
            row_number_item.setTextAlignment(
                Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter
            )
            self.mapping_table.setItem(row_idx, 0, row_number_item)

            self._template_row_to_table_index[event_row.row_index] = row_idx

            label_item = QTableWidgetItem(event_row.label)
            label_item.setFlags(label_item.flags() & ~no_edit)
            self.mapping_table.setItem(row_idx, 1, label_item)

            combo = QComboBox()
            combo.setEditable(False)
            combo.setSizeAdjustPolicy(
                QComboBox.SizeAdjustPolicy.AdjustToMinimumContentsLengthWithIcon
            )
            combo.setMinimumContentsLength(14)
            combo.setStyleSheet(_get_cell_combobox_stylesheet())
            combo.addItem("<unmapped>", None)
            for event in wiz.session_events:
                combo.addItem(event.combo_text, event.index)

            assignment = wiz.row_assignments.get(event_row.row_index)
            if assignment is not None:
                idx = combo.findData(assignment)
                if idx >= 0:
                    combo.setCurrentIndex(idx)

            combo.currentIndexChanged.connect(
                self._make_row_selection_handler(event_row.row_index, combo)
            )
            self.mapping_table.setCellWidget(row_idx, 2, combo)
            self._event_row_widgets[event_row.row_index] = combo

            # col 3: mapped value
            value_item = QTableWidgetItem("")
            value_item.setFlags(value_item.flags() & ~no_edit)
            value_item.setTextAlignment(
                Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter
            )
            self.mapping_table.setItem(row_idx, 3, value_item)
            self._value_items[event_row.row_index] = value_item

            # col 4: current template value
            current_item = QTableWidgetItem("")
            current_item.setFlags(current_item.flags() & ~no_edit)
            current_item.setTextAlignment(
                Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter
            )
            muted_color = get_semantic_colors()["muted"]
            current_item.setForeground(QBrush(QColor(muted_color)))
            self.mapping_table.setItem(row_idx, 4, current_item)
            self._current_value_items[event_row.row_index] = current_item

            # col 5: status indicator
            status_item = QTableWidgetItem("")
            status_item.setFlags(status_item.flags() & ~no_edit)
            status_item.setTextAlignment(
                Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter
            )
            self.mapping_table.setItem(row_idx, 5, status_item)
            self._status_items[event_row.row_index] = status_item

        self._refresh_current_value_column()
        self._refresh_value_column()
        self._refresh_status_icons()

    # --------------------------------------------------
    def _refresh_value_column(self) -> None:
        wiz = self._wizard()
        measurement = wiz.current_measurement
        for row_index, item in self._value_items.items():
            assignment = wiz.row_assignments.get(row_index)
            text = ""
            if assignment is not None:
                value = wiz.value_for_event(assignment, measurement)
                if pd.isna(value):
                    text = "—"
                elif isinstance(value, Real):
                    text = f"{value:.2f}"
                else:
                    text = str(value)
            item.setText(text)

    # --------------------------------------------------
    def _refresh_current_value_column(self) -> None:
        """Show the value currently in the Excel template cell."""
        wiz = self._wizard()
        date_col = wiz.active_date_column
        for row_index, item in self._current_value_items.items():
            text = ""
            if date_col is not None and wiz.ws is not None:
                cell_val = wiz.ws.cell(
                    row=row_index, column=date_col.column_index
                ).value
                if cell_val is not None:
                    if isinstance(cell_val, (int, float)) and not isinstance(cell_val, bool):
                        text = f"{cell_val:.2f}"
                    else:
                        text = str(cell_val)
            item.setText(text)

    # --------------------------------------------------
    def _refresh_status_icons(self) -> None:
        wiz = self._wizard()
        assignments = Counter(
            a for a in wiz.row_assignments.values() if a is not None
        )
        duplicates = {idx for idx, count in assignments.items() if count > 1}

        colors = get_semantic_colors()
        table_bg = theme.CURRENT_THEME.get("table_bg", "#020617")
        col_count = self.mapping_table.columnCount()

        # Subtle row tint colors (low-alpha blends)
        mapped_bg = QColor(colors["success"])
        mapped_bg.setAlpha(25)
        warning_bg = QColor(colors["warning"])
        warning_bg.setAlpha(25)
        unmapped_bg = QColor(table_bg)  # no tint

        for row_index, status_item in self._status_items.items():
            table_row = self._template_row_to_table_index.get(row_index)
            if table_row is None:
                continue
            assignment = wiz.row_assignments.get(row_index)

            # -- status indicator --
            font = status_item.font()
            font.setPointSize(14)
            font.setBold(False)
            if assignment is None:
                status_item.setText("○")
                status_item.setToolTip("Not mapped")
                status_item.setForeground(QBrush(QColor(colors["error"])))
                row_bg = unmapped_bg
            elif assignment in duplicates:
                status_item.setText("⚠")
                status_item.setToolTip("Event reused on multiple rows")
                status_item.setForeground(QBrush(QColor(colors["warning"])))
                row_bg = warning_bg
            else:
                status_item.setText("✓")
                status_item.setToolTip("Mapped")
                status_item.setForeground(QBrush(QColor(colors["success"])))
                font.setBold(True)
                row_bg = mapped_bg
            status_item.setFont(font)

            # -- row background tint --
            brush = QBrush(row_bg)
            for col in range(col_count):
                cell = self.mapping_table.item(table_row, col)
                if cell is not None:
                    cell.setBackground(brush)

    # --------------------------------------------------
    def _apply_assignment(self, row_index: int, event_index: int | None) -> None:
        wiz = self._wizard()
        wiz.update_row_assignment(row_index, event_index)
        self._refresh_value_column()
        self._refresh_status_icons()
        self._update_status_banner()

    # --------------------------------------------------
    def _update_status_banner(self) -> None:
        wiz = self._wizard()
        colors = get_semantic_colors()
        if getattr(wiz, "manual_date_selection_required", False):
            self.info_label.setText(
                "Multiple date columns detected. Pick the active date column before continuing."
            )
            self.info_label.setStyleSheet(f"color: {colors['warning']};")
            return
        unmapped = sum(1 for value in wiz.row_assignments.values() if value is None)
        if unmapped:
            self.info_label.setText(
                f"{unmapped} event row(s) are still unmapped. Only mapped rows will be written."
            )
            self.info_label.setStyleSheet(f"color: {colors['warning']};")
            self.select_unmapped_btn.setVisible(True)
        else:
            self.info_label.setText(
                "Review the mappings below. You can override any row before saving."
            )
            self.info_label.setStyleSheet(f"color: {colors['muted']};")
            self.select_unmapped_btn.setVisible(False)

    # --------------------------------------------------
    def _on_measurement_changed(self, value: str) -> None:
        wiz = self._wizard()
        wiz.current_measurement = value
        self._refresh_value_column()
        self._refresh_current_value_column()
        self._refresh_status_icons()
        self._update_status_banner()

    # --------------------------------------------------
    def _on_redetect(self) -> None:
        wiz = self._wizard()
        if not wiz.prepare_layout(auto=True, force=True):
            colors = get_semantic_colors()
            self.info_label.setText(wiz.layout_error or "Could not re-run detection.")
            self.info_label.setStyleSheet(f"color: {colors['error']};")
            return
        self._populate_date_options()
        self._rebuild_mapping_table()
        self._update_status_banner()

    # --------------------------------------------------
    def _on_date_changed(self, index: int) -> None:
        if index < 0:
            return
        option = self.pick_date_combo.itemData(index)
        if not isinstance(option, DateColumnOption):
            return
        wiz = self._wizard()
        wiz.set_active_date_column(option)
        wiz.ensure_active_date_value(self)
        self._refresh_current_value_column()
        self._refresh_status_icons()

    # --------------------------------------------------
    def _make_row_selection_handler(self, row_index: int, combo: QComboBox):
        def handler() -> None:
            value = combo.currentData()
            self._apply_assignment(row_index, value)

        return handler

    # --------------------------------------------------
    def _select_all_unmapped(self) -> None:
        for _row_index, combo in self._event_row_widgets.items():
            if combo.currentData() is None:
                combo.showPopup()
                break

    # --------------------------------------------------
    def validatePage(self) -> bool:
        wiz = self._wizard()
        wiz.persist_assignments()
        return True

    # --------------------------------------------------
    def isComplete(self) -> bool:
        return True


class PreviewPage(WizardPageBase):
    """Final page showing a preview and allowing export."""

    def __init__(self) -> None:
        super().__init__()
        self.setTitle("Step 4: Preview & Save")
        layout = QVBoxLayout(self)
        self.preview_view = QTableView()
        self.btn_save = QPushButton("Update Template")
        self.btn_save.setMinimumHeight(BUTTON_HEIGHT)
        self.btn_save.setObjectName("PrimaryButton")  # Make it primary
        self.btn_save.setStyleSheet(get_modern_button_stylesheet())
        self.btn_save.clicked.connect(self.save_file)
        layout.addWidget(self.preview_view, 1)
        layout.addWidget(self.btn_save)

    # ------------------------------------------------------
    def initializePage(self) -> None:
        wiz = self._wizard()
        if wiz is None:
            return

        wiz.apply_mapping()
        preview_df = wiz.get_preview_dataframe()
        self.preview_view.setModel(PandasModel(preview_df))
        self.preview_view.setStyleSheet(get_modern_table_stylesheet())
        has_mappings = any(value is not None for value in wiz.row_assignments.values())
        self.btn_save.setEnabled(has_mappings)

    # ------------------------------------------------------
    def save_file(self) -> None:
        template_path = self.field("templatePath")
        if not template_path:
            QMessageBox.warning(self, "Error", "Template path not set.")
            return

        target_path = Path(template_path)
        if not target_path.exists():
            QMessageBox.warning(
                self,
                "Missing File",
                "The original template file is no longer available."
                " Please choose a new template before saving.",
            )
            return

        # Pre-export validation
        wiz = self._wizard()
        warnings = self._validate_before_export(wiz)
        if warnings:
            warning_text = "\n".join(f"• {w}" for w in warnings)
            proceed = QMessageBox.warning(
                self,
                "Mapping Warnings",
                f"{warning_text}\n\nContinue anyway?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            )
            if proceed != QMessageBox.StandardButton.Yes:
                return

        confirm = QMessageBox.question(
            self,
            "Update Template",
            f"This will overwrite {target_path.name} with the mapped values. Continue?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if confirm != QMessageBox.StandardButton.Yes:
            return

        wiz.apply_mapping()

        try:
            save_workbook(wiz.wb, str(target_path))
        except PermissionError:  # pragma: no cover - GUI feedback
            QMessageBox.warning(
                self,
                "File In Use",
                (
                    f"Could not update {target_path.name} because it appears to be open in Excel"
                    " or another program.\n\nPlease close the file and try again."
                ),
            )
            return
        except Exception as exc:  # pragma: no cover - GUI feedback
            QMessageBox.critical(self, "Save Failed", str(exc))
            return

        # Persist mapping corrections for future sessions
        wiz.save_mapping_history()

        QMessageBox.information(self, "Template Updated", f"Mappings written to {target_path}")
        self.completeChanged.emit()

    # ------------------------------------------------------
    @staticmethod
    def _validate_before_export(wiz: "ExcelMapWizard") -> list[str]:
        """Return a list of human-readable warnings (empty = all good)."""
        warnings: list[str] = []

        # Unmapped rows
        unmapped = sum(1 for v in wiz.row_assignments.values() if v is None)
        if unmapped:
            warnings.append(
                f"{unmapped} template row(s) have no mapped event and will be skipped."
            )

        # Duplicate event assignments
        assignments = Counter(a for a in wiz.row_assignments.values() if a is not None)
        dupes = {idx for idx, count in assignments.items() if count > 1}
        if dupes:
            event_by_index = {e.index: e for e in wiz.session_events}
            dupe_labels = [event_by_index[i].label for i in dupes if i in event_by_index]
            warnings.append(
                f"Event(s) mapped to multiple rows: {', '.join(dupe_labels)}"
            )

        return warnings

    # ------------------------------------------------------
    def isComplete(self) -> bool:
        return self._wizard().wb is not None

    def apply_theme(self) -> None:
        self.btn_save.setStyleSheet(get_modern_button_stylesheet())
        self.preview_view.setStyleSheet(get_modern_table_stylesheet())


# ---------------------------------------------------------------------------
# Main wizard class
# ---------------------------------------------------------------------------


class ExcelMapWizard(QWizard):
    """Wizard dialog used to map events to Excel templates."""

    MAX_PREVIEW_COLUMNS = 6

    def __init__(self, parent=None, events_df: pd.DataFrame | None = None) -> None:
        super().__init__(parent)
        self.setWindowTitle("Map Events to Excel")
        self.setWizardStyle(QWizard.WizardStyle.ModernStyle)
        self.setOption(QWizard.WizardOption.HaveFinishButtonOnEarlyPages)
        self.setMinimumSize(1100, 720)
        self.resize(1280, 780)

        # Apply modern dialog styling
        combined_stylesheet = (
            get_modern_button_stylesheet() + "\n" + get_modern_combobox_stylesheet()
        )
        self.setStyleSheet(combined_stylesheet)

        # Workbook/session state
        self.wb = None
        self.ws = None
        self.selected_sheet_name: str | None = None
        self.eventsDF: pd.DataFrame | None = None
        self.events_source: str | None = None

        # Template metadata (from VBA macros or auto-detection)
        self.template_metadata: TemplateMetadata | None = None

        # Derived session data
        self.session_events: list[SessionEventInfo] = []
        self.measurement_columns: list[str] = []
        self.current_measurement: str | None = None

        # Template layout metadata
        self.values_block: tuple[int, int, int, int] | None = None
        self.date_row_index: int | None = None
        self.date_columns: list[DateColumnOption] = []
        self.active_date_column: DateColumnOption | None = None
        self.manual_date_selection_required: bool = False
        self.pending_new_date: bool = False
        self.date_columns_bounds: tuple[int, int] = (0, 0)

        # Row mapping state
        self.event_rows: list[EventRowInfo] = []
        self.row_assignments: dict[int, int | None] = {}

        # Cell history for safe replay
        self._original_values: dict[str, Any] = {}
        self._mapped_cells: set[str] = set()

        # Layout detection cache
        self._layout_ready = False
        self.layout_error = ""

        self.addPage(TemplatePage())
        self.addPage(RowMappingPage())
        self.addPage(PreviewPage())
        self.apply_theme()

        if events_df is not None:
            self.set_events_dataframe(events_df, source="session")

    def apply_theme(self) -> None:
        """Re-apply dialog styles after a theme change."""
        combined_stylesheet = (
            get_modern_button_stylesheet() + "\n" + get_modern_combobox_stylesheet()
        )
        self.setStyleSheet(combined_stylesheet)

        for page in self.findChildren(QWizardPage):
            apply_method = getattr(page, "apply_theme", None)
            if callable(apply_method):
                with contextlib.suppress(Exception):
                    apply_method()

        for table in self.findChildren((QTableWidget, QTableView)):
            apply_method = getattr(table, "apply_theme", None)
            if callable(apply_method):
                with contextlib.suppress(Exception):
                    apply_method()
            else:
                with contextlib.suppress(Exception):
                    table.setStyleSheet(get_modern_table_stylesheet())

        self.update()

    # --------------------------------------------------
    def set_events_dataframe(self, df: pd.DataFrame, *, source: str = "session") -> None:
        self.eventsDF = df.reset_index(drop=True)
        self.events_source = source

        numeric_cols: list[str] = []
        for col in self.eventsDF.columns:
            if col == "Event":
                continue
            series = self.eventsDF[col]
            if pd.api.types.is_numeric_dtype(series):
                numeric_cols.append(col)

        if not numeric_cols:
            raise ValueError("Event data must include at least one numeric measurement column.")

        preferred_order = ["ID (µm)", "Inner Diameter (µm)", "ID"]
        for pref in preferred_order:
            if pref in numeric_cols:
                self.current_measurement = pref
                break
        else:
            if self.current_measurement not in numeric_cols:
                self.current_measurement = numeric_cols[0]

        self.measurement_columns = numeric_cols
        self.session_events = []
        time_series = self.eventsDF["Time (s)"] if "Time (s)" in self.eventsDF.columns else None
        for idx, row in self.eventsDF.iterrows():
            label = str(row.get("Event", f"Event {idx + 1}"))
            time_val = (
                float(row["Time (s)"])
                if time_series is not None and not pd.isna(row["Time (s)"])
                else None
            )
            values = {col: row[col] for col in numeric_cols}
            self.session_events.append(
                SessionEventInfo(index=idx, label=label, time_value=time_val, values=values)
            )

        self._layout_ready = False
        self.row_assignments = {}

    # --------------------------------------------------
    def clear_mapped_cells(self) -> None:
        if not self.ws:
            self._mapped_cells.clear()
            return
        for cell in list(self._mapped_cells):
            if cell in self._original_values:
                self.ws[cell] = self._original_values[cell]
        self._mapped_cells.clear()

    # --------------------------------------------------
    def _remember_original(self, cell_address: str) -> None:
        if self.ws is None:
            return
        if cell_address not in self._original_values:
            self._original_values[cell_address] = self.ws[cell_address].value

    # --------------------------------------------------
    def reset_mapping_state(self) -> None:
        self.clear_mapped_cells()
        self._original_values = {}
        self._layout_ready = False
        self.layout_error = ""
        self.date_columns = []
        self.active_date_column = None
        self.manual_date_selection_required = False
        self.pending_new_date = False
        self.event_rows = []
        self.row_assignments = {}

    # --------------------------------------------------
    def _get_defined_range(self, name: str) -> str | None:
        if self.wb is None or self.ws is None:
            return None
        defined = self.wb.defined_names.get(name)
        if not defined:
            return None
        destinations = list(defined.destinations)

        # First, try to find a range specifically for the selected sheet
        for sheet_name, coord in destinations:
            if sheet_name == self.ws.title:
                return coord

        # Fallback: if there's only one destination and it references a different sheet,
        # use that range anyway (assumes same structure across sheets)
        if destinations and len(destinations) == 1:
            sheet_name, coord = destinations[0]
            # Use the same coordinate range on the selected sheet
            return coord

        return None

    # --------------------------------------------------
    def _resolve_ranges(self) -> None:
        """
        Resolve template ranges from metadata or named ranges.

        Priority:
        1. Template metadata (from VBA macros)
        2. Named ranges (VASO_DATES_ROW, VASO_VALUES_BLOCK)
        3. Raise error if neither found
        """
        # Try metadata first
        if self.template_metadata:
            self.date_row_index = self.template_metadata.date_row

            # Build values block from metadata event rows
            if self.template_metadata.event_rows:
                min_row = min(row.row for row in self.template_metadata.event_rows)
                max_row = max(row.row for row in self.template_metadata.event_rows)
            elif self.template_metadata.event_rows_start and self.template_metadata.event_rows_end:
                min_row = self.template_metadata.event_rows_start
                max_row = self.template_metadata.event_rows_end
            else:
                raise ValueError("Template metadata has no event row information")

            # Date column bounds
            if self.template_metadata.date_columns:
                min_col = min(col.column for col in self.template_metadata.date_columns)
                max_col = max(col.column for col in self.template_metadata.date_columns)
            elif (
                self.template_metadata.date_columns_start
                and self.template_metadata.date_columns_end
            ):
                min_col = self.template_metadata.date_columns_start
                max_col = self.template_metadata.date_columns_end
            else:
                # Default to B:Z
                min_col, max_col = 2, 26

            self.values_block = (min_row, max_row, min_col, max_col)
            self.date_columns_bounds = (min_col, max_col)
            return

        # Fallback to named ranges
        date_range = self._get_defined_range("VASO_DATES_ROW")
        if not date_range:
            raise ValueError(
                "Workbook is missing defined name 'VASO_DATES_ROW'.\n\n"
                "Either add VasoAnalyzer metadata (see Excel Template Setup guide)\n"
                "or define named ranges VASO_DATES_ROW and VASO_VALUES_BLOCK."
            )
        d_min_col, d_min_row, d_max_col, d_max_row = range_boundaries(date_range)
        if d_min_row != d_max_row:
            raise ValueError("VASO_DATES_ROW must refer to a single row.")
        self.date_row_index = d_min_row

        values_range = self._get_defined_range("VASO_VALUES_BLOCK")
        if not values_range:
            raise ValueError(
                "Workbook is missing defined name 'VASO_VALUES_BLOCK'.\n\n"
                "Either add VasoAnalyzer metadata (see Excel Template Setup guide)\n"
                "or define named ranges VASO_DATES_ROW and VASO_VALUES_BLOCK."
            )
        v_min_col, v_min_row, v_max_col, v_max_row = range_boundaries(values_range)
        self.values_block = (v_min_row, v_max_row, v_min_col, v_max_col)
        self.date_columns_bounds = (d_min_col, d_max_col)

    # --------------------------------------------------
    @staticmethod
    def _has_fill(cell) -> bool:
        fill = getattr(cell, "fill", None)
        if not fill:
            return False
        pattern = getattr(fill, "patternType", None)
        return not (not pattern or pattern.lower() == "none")

    # --------------------------------------------------
    def _extract_event_rows(self) -> None:
        """
        Extract event rows from template.

        Uses metadata if available, otherwise scans the values block.
        """
        if self.ws is None or self.values_block is None:
            self.event_rows = []
            return

        # Use metadata if available
        if self.template_metadata and self.template_metadata.event_rows:
            rows: list[EventRowInfo] = []
            for meta_row in self.template_metadata.event_rows:
                cell = self.ws.cell(row=meta_row.row, column=self.template_metadata.label_column)
                rows.append(
                    EventRowInfo(
                        row_index=meta_row.row,
                        label=meta_row.label,
                        is_header=meta_row.is_header,
                        label_cell=cell.coordinate,
                    )
                )
            self.event_rows = rows
        else:
            # Fallback: scan values block
            min_row, max_row, _, _ = self.values_block
            scanned_rows: list[EventRowInfo] = []
            for row_idx in range(min_row, max_row + 1):
                cell = self.ws.cell(row=row_idx, column=1)
                value = cell.value
                label = str(value).strip() if value not in (None, "") else ""
                if not label:
                    continue
                is_header = bool(
                    getattr(cell, "font", None) and getattr(cell.font, "bold", False)
                ) and self._has_fill(cell)
                scanned_rows.append(
                    EventRowInfo(
                        row_index=row_idx,
                        label=label,
                        is_header=is_header,
                        label_cell=cell.coordinate,
                    )
                )
            self.event_rows = scanned_rows

        valid_rows = {row.row_index for row in self.event_rows if not row.is_header}
        self.row_assignments = {
            row_idx: self.row_assignments.get(row_idx) for row_idx in valid_rows
        }

    # --------------------------------------------------
    def _build_date_options(self) -> None:
        """
        Build date column options.

        Uses metadata if available, otherwise scans date row.
        """
        self.date_columns = []
        if self.ws is None or self.values_block is None or self.date_row_index is None:
            return

        # Use metadata if available
        if self.template_metadata and self.template_metadata.date_columns:
            for meta_col in self.template_metadata.date_columns:
                cell = self.ws.cell(row=self.date_row_index, column=meta_col.column)
                # Re-count empty slots (metadata might be stale)
                empty_slots = 0
                for event_row in self.event_rows:
                    if event_row.is_header:
                        continue
                    target_cell = self.ws.cell(row=event_row.row_index, column=meta_col.column)
                    if target_cell.value in (None, ""):
                        empty_slots += 1
                option = DateColumnOption(
                    column_index=meta_col.column,
                    cell_address=cell.coordinate,
                    value=cell.value,  # Use current value, not cached
                    empty_slots=empty_slots,
                )
                self.date_columns.append(option)
        else:
            # Fallback: scan date row
            min_row, max_row, min_col, max_col = self.values_block
            d_min_col, d_max_col = self.date_columns_bounds

            for col_idx in range(d_min_col, d_max_col + 1):
                cell = self.ws.cell(row=self.date_row_index, column=col_idx)
                value = cell.value
                empty_slots = 0
                for event_row in self.event_rows:
                    if event_row.is_header:
                        continue
                    if not (min_col <= col_idx <= max_col):
                        continue
                    target_cell = self.ws.cell(row=event_row.row_index, column=col_idx)
                    if target_cell.value in (None, ""):
                        empty_slots += 1
                option = DateColumnOption(
                    column_index=col_idx,
                    cell_address=cell.coordinate,
                    value=value,
                    empty_slots=empty_slots,
                )
                self.date_columns.append(option)

    # --------------------------------------------------
    @staticmethod
    def _is_valid_date(value: Any) -> bool:
        if value in (None, ""):
            return False
        try:
            pd.to_datetime(value)
            return True
        except (ValueError, TypeError, pd.errors.OutOfBoundsDatetime):
            return False

    # --------------------------------------------------
    def auto_select_date_column(self) -> None:
        self.manual_date_selection_required = False
        self.pending_new_date = False
        if not self.date_columns:
            self.active_date_column = None
            return

        if self.active_date_column:
            for option in self.date_columns:
                if option.column_index == self.active_date_column.column_index:
                    self.active_date_column = option
                    break

        valid_dates = [opt for opt in self.date_columns if self._is_valid_date(opt.value)]
        if len(valid_dates) == 1:
            self.active_date_column = valid_dates[0]
            return
        if len(valid_dates) > 1:
            empties = [opt for opt in valid_dates if opt.empty_slots > 0]
            if empties:
                self.active_date_column = max(empties, key=lambda opt: opt.column_index)
            else:
                self.active_date_column = max(valid_dates, key=lambda opt: opt.column_index)
                self.manual_date_selection_required = True
            return

        empty_cells = [opt for opt in self.date_columns if opt.value in (None, "")]
        if empty_cells:
            self.active_date_column = empty_cells[0]
            self.pending_new_date = True
            return

        # Fallback to last column if everything contains non-date data
        self.active_date_column = self.date_columns[-1]
        self.manual_date_selection_required = True

    # --------------------------------------------------
    def ensure_active_date_value(self, parent) -> None:
        if self.ws is None or not self.active_date_column:
            return
        cell = self.ws[self.active_date_column.cell_address]
        if cell.value not in (None, "") and not self.pending_new_date:
            return

        prompt = "Enter the column label (e.g., experiment date) for this mapping:"
        text, ok = QInputDialog.getText(parent, "Set Date Label", prompt)
        if not ok or not text.strip():
            return

        self._remember_original(self.active_date_column.cell_address)
        value = text.strip()
        cell.value = value
        self.active_date_column.value = value
        self.active_date_column.is_new = True
        self._mapped_cells.add(self.active_date_column.cell_address)
        self.pending_new_date = False
        self._update_date_option_value(self.active_date_column.column_index, value)

    # --------------------------------------------------
    def _update_date_option_value(self, column_index: int, value: Any) -> None:
        for option in self.date_columns:
            if option.column_index == column_index:
                option.value = value
                break

    # --------------------------------------------------
    def set_active_date_column(self, option: DateColumnOption) -> None:
        self.active_date_column = option
        self.manual_date_selection_required = False
        self.auto_assign_rows(force=False)

    # --------------------------------------------------
    def update_row_assignment(self, row_index: int, event_index: int | None) -> None:
        if event_index is None:
            self.row_assignments[row_index] = None
        else:
            self.row_assignments[row_index] = int(event_index)

    # --------------------------------------------------
    @staticmethod
    def _normalize_numeric(value: Any) -> Any:
        try:
            return round(float(value), 4)
        except (ValueError, TypeError):
            return str(value).strip()

    # --------------------------------------------------
    def _load_existing_assignments_from_sheet(self) -> None:
        if self.ws is None or self.active_date_column is None or not self.current_measurement:
            return
        col_idx = self.active_date_column.column_index
        measurement = self.current_measurement
        value_map: dict[Any, list[int]] = {}
        for event in self.session_events:
            value = event.values.get(measurement)
            if pd.isna(value):
                continue
            key = self._normalize_numeric(value)
            value_map.setdefault(key, []).append(event.index)

        for row in self.event_rows:
            if row.is_header:
                continue
            if (
                row.row_index not in self.row_assignments
                or self.row_assignments[row.row_index] is not None
            ):
                continue
            cell_value = self.ws.cell(row=row.row_index, column=col_idx).value
            if cell_value in (None, ""):
                continue
            key = self._normalize_numeric(cell_value)
            matches = value_map.get(key)
            if matches:
                self.row_assignments[row.row_index] = matches[0]

    # --------------------------------------------------
    def auto_assign_rows(self, force: bool = False) -> None:
        valid_rows = [row.row_index for row in self.event_rows if not row.is_header]
        if force or not self.row_assignments:
            self.row_assignments = {row_idx: None for row_idx in valid_rows}
        else:
            self.row_assignments = {
                row_idx: self.row_assignments.get(row_idx) for row_idx in valid_rows
            }

        if self.active_date_column:
            self._load_existing_assignments_from_sheet()

        # Pass 0: apply saved mapping history (user corrections from previous sessions)
        self._apply_mapping_history()

        # Pass 1: exact case-insensitive match
        label_map: dict[str, deque[int]] = {}
        for event in self.session_events:
            key = event.label.strip().lower()
            label_map.setdefault(key, deque()).append(event.index)

        for row in self.event_rows:
            if row.is_header:
                continue
            if self.row_assignments.get(row.row_index) is not None:
                continue
            key = row.label.strip().lower()
            queue = label_map.get(key)
            if queue:
                self.row_assignments[row.row_index] = queue.popleft()

        # Pass 2: normalized match — handles µ→u, punctuation, and spacing
        # differences between session event labels and template row labels.
        norm_map: dict[str, deque[int]] = {}
        for event in self.session_events:
            key = normalize_label(event.label)
            norm_map.setdefault(key, deque()).append(event.index)

        for row in self.event_rows:
            if row.is_header:
                continue
            if self.row_assignments.get(row.row_index) is not None:
                continue  # already matched in pass 1
            key = normalize_label(row.label)
            queue = norm_map.get(key)
            if queue:
                self.row_assignments[row.row_index] = queue.popleft()

        # Pass 3: fuzzy match — substring containment + token-overlap scoring
        # for rows that still have no assignment after exact/normalised passes.
        assigned_indices: set[int] = {
            idx for idx in self.row_assignments.values() if idx is not None
        }
        remaining_labels = [
            e.label
            for e in self.session_events
            if e.index not in assigned_indices
        ]
        for row in self.event_rows:
            if row.is_header:
                continue
            if self.row_assignments.get(row.row_index) is not None:
                continue
            match_label = best_match(row.label, remaining_labels)
            if match_label is not None:
                # Find the event index for this label
                for event in self.session_events:
                    if event.label == match_label and event.index not in assigned_indices:
                        self.row_assignments[row.row_index] = event.index
                        assigned_indices.add(event.index)
                        remaining_labels.remove(match_label)
                        break

    # --------------------------------------------------
    # Mapping persistence — remember user corrections
    # --------------------------------------------------

    def _template_fingerprint(self) -> str:
        """Return a stable hash identifying this template by its row labels."""
        labels = sorted(
            normalize_label(r.label) for r in self.event_rows if not r.is_header
        )
        raw = "|".join(labels)
        return hashlib.sha256(raw.encode()).hexdigest()[:16]

    def _load_mapping_history(self) -> dict[str, str]:
        """Load previously saved template_label → session_label map."""
        settings = QSettings("TykockiLab", "VasoAnalyzer")
        key = f"excel_mapping_history/{self._template_fingerprint()}"
        data = settings.value(key, "", type=str)
        if data:
            try:
                return json.loads(data)
            except (json.JSONDecodeError, TypeError):
                pass
        return {}

    def save_mapping_history(self) -> None:
        """Persist current row_assignments as template_label → session_label."""
        history: dict[str, str] = {}
        event_by_index = {e.index: e for e in self.session_events}
        for row in self.event_rows:
            if row.is_header:
                continue
            ev_idx = self.row_assignments.get(row.row_index)
            if ev_idx is not None and ev_idx in event_by_index:
                history[row.label] = event_by_index[ev_idx].label
        if history:
            settings = QSettings("TykockiLab", "VasoAnalyzer")
            key = f"excel_mapping_history/{self._template_fingerprint()}"
            settings.setValue(key, json.dumps(history))

    def _apply_mapping_history(self) -> None:
        """Apply saved mappings as Pass 0 (before fuzzy matching)."""
        history = self._load_mapping_history()
        if not history:
            return
        # Build label→event index lookup
        label_to_events: dict[str, list[int]] = {}
        for event in self.session_events:
            label_to_events.setdefault(event.label, []).append(event.index)

        assigned: set[int] = set()
        for row in self.event_rows:
            if row.is_header:
                continue
            if self.row_assignments.get(row.row_index) is not None:
                continue
            saved_label = history.get(row.label)
            if saved_label and saved_label in label_to_events:
                candidates = label_to_events[saved_label]
                for idx in candidates:
                    if idx not in assigned:
                        self.row_assignments[row.row_index] = idx
                        assigned.add(idx)
                        break

    # --------------------------------------------------
    def value_for_event(self, event_index: int, measurement: str | None) -> Any:
        if measurement is None:
            return float("nan")
        if event_index < 0 or event_index >= len(self.session_events):
            return float("nan")
        return self.session_events[event_index].values.get(measurement, float("nan"))

    # --------------------------------------------------
    def prepare_layout(self, auto: bool = False, force: bool = False) -> bool:
        if force:
            self._layout_ready = False

        if self._layout_ready and not force:
            if auto:
                self.auto_select_date_column()
                self.auto_assign_rows(force=False)
            return True

        if self.wb is None or self.ws is None or self.eventsDF is None:
            self.layout_error = "Load an Excel template and event data first."
            log.warning(
                f"[Wizard] prepare_layout blocked: wb={self.wb is not None}, "
                f"ws={self.ws is not None}, eventsDF={self.eventsDF is not None}"
            )
            return False

        log.info(
            f"[Wizard] prepare_layout starting. "
            f"Sheet: {getattr(self.ws, 'title', '?')!r}, "
            f"metadata: {self.template_metadata.source if self.template_metadata else 'None'}, "
            f"events: {len(self.eventsDF)}"
        )

        try:
            self._resolve_ranges()
            log.info(
                f"[Wizard] Ranges resolved: values_block={self.values_block}, "
                f"date_row={self.date_row_index}, date_bounds={self.date_columns_bounds}"
            )
            self._extract_event_rows()
            log.info(
                f"[Wizard] Event rows extracted: {len(self.event_rows)} total, "
                f"{sum(1 for r in self.event_rows if not r.is_header)} mappable"
            )
            self._build_date_options()
            log.info(
                f"[Wizard] Date columns: {[c.letter for c in self.date_columns]}"
            )
            self.auto_select_date_column()
            self.auto_assign_rows(force=True)
            matched = sum(1 for v in self.row_assignments.values() if v is not None)
            log.info(
                f"[Wizard] Auto-assign done: {matched}/{len(self.row_assignments)} rows matched"
            )
            self._layout_ready = True
            self.layout_error = ""
        except Exception as exc:  # pragma: no cover - GUI feedback
            log.error(f"[Wizard] prepare_layout failed: {exc}", exc_info=True)
            self.layout_error = str(exc)
            self._layout_ready = False
            return False

        return True

    # --------------------------------------------------
    def preview_template_data(self, limit: int = 30) -> list[dict[str, Any]]:
        if self.ws is None or self.values_block is None:
            return []

        min_row, max_row, min_col, max_col = self.values_block
        active_col = self.active_date_column.column_index if self.active_date_column else None

        cols = list(range(min_col, max_col + 1))
        if len(cols) > self.MAX_PREVIEW_COLUMNS:
            cols = cols[: self.MAX_PREVIEW_COLUMNS]
            if active_col and active_col not in cols:
                cols.append(active_col)
        cols = sorted(set(cols))

        data: list[dict[str, Any]] = []
        for row_idx in range(min_row, max_row + 1):
            if len(data) >= limit:
                break
            label_cell = self.ws.cell(row=row_idx, column=1)
            label = label_cell.value if label_cell.value not in (None, "") else ""
            is_header = False
            is_event = False
            for event_row in self.event_rows:
                if event_row.row_index == row_idx:
                    is_header = event_row.is_header
                    is_event = not event_row.is_header
                    break
            row_data: dict[str, Any] = {
                "Row": row_idx,
                "Label": label,
                "_is_header": is_header,
                "_is_event": is_event,
                "_active_col": active_col,
            }
            for col_idx in cols:
                letter = get_column_letter(col_idx)
                value = self.ws.cell(row=row_idx, column=col_idx).value
                row_data[letter] = value
            data.append(row_data)
        return data

    # --------------------------------------------------
    def apply_mapping(self) -> list[tuple[int, str, Any]]:
        self.clear_mapped_cells()
        results: list[tuple[int, str, Any]] = []

        if (
            self.ws is None
            or self.values_block is None
            or self.active_date_column is None
            or not self.current_measurement
        ):
            return results

        min_row, max_row, min_col, max_col = self.values_block
        if not (min_col <= self.active_date_column.column_index <= max_col):
            return results

        date_cell = self.ws[self.active_date_column.cell_address]
        if self.active_date_column.value not in (None, ""):
            self._remember_original(self.active_date_column.cell_address)
            date_cell.value = self.active_date_column.value
            self._mapped_cells.add(self.active_date_column.cell_address)

        measurement = self.current_measurement
        for row in self.event_rows:
            if row.is_header:
                continue
            assignment = self.row_assignments.get(row.row_index)
            if assignment is None:
                continue
            if not (min_row <= row.row_index <= max_row):
                continue
            cell_address = f"{self.active_date_column.letter}{row.row_index}"
            self._remember_original(cell_address)
            value = self.value_for_event(assignment, measurement)
            cell_value = None if pd.isna(value) else value
            self.ws[cell_address] = cell_value
            self._mapped_cells.add(cell_address)
            results.append((row.row_index, row.label, cell_value))

        return results

    # --------------------------------------------------
    def persist_assignments(self) -> None:
        # Placeholder for future persistence (e.g., to workbook metadata)
        pass

    # --------------------------------------------------
    def get_preview_dataframe(self, limit: int = 25) -> pd.DataFrame:
        if self.ws is None or self.active_date_column is None or not self.current_measurement:
            return pd.DataFrame()

        rows: list[dict[str, Any]] = []
        col_idx = self.active_date_column.column_index
        measurement = self.current_measurement

        for row in self.event_rows:
            if row.is_header:
                continue
            cell_value = self.ws.cell(row=row.row_index, column=col_idx).value
            rows.append({"Row": row.row_index, "Label": row.label, measurement: cell_value})
            if len(rows) >= limit:
                break

        return pd.DataFrame(rows)
