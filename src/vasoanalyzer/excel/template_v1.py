"""Inspection and validation helpers for the VasoAnalyzer standard template (v1)."""

from __future__ import annotations

from dataclasses import dataclass
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

SIGNATURE_NAME = "VA_TEMPLATE_SIGNATURE"
SIGNATURE_VALUE = "VasoAnalyzerTemplate:v1"
BLOCK_PREFIX = "VA_Block_"


@dataclass(frozen=True)
class ExcelBlock:
    block_id: str
    sheet_name: str
    table_name: str
    metric_col: str
    replicate_cols: list[str]
    summary_cols: list[str]
    title: str


@dataclass(frozen=True)
class TemplateInspection:
    signature_ok: bool
    version: str
    blocks: list[ExcelBlock]
    warnings: list[str]


def _resolve_defined_name_value(workbook, name: str) -> str | None:
    defined = workbook.defined_names.get(name)
    if defined is None:
        return None
    if isinstance(defined, list):
        defined = defined[0]
    try:
        for title, coord in defined.destinations:
            ws = workbook[title]
            return ws[coord].value
    except Exception:
        pass
    attr_text = getattr(defined, "attr_text", None)
    if isinstance(attr_text, str):
        return attr_text.strip().strip('"')
    return None


def _merged_intersects(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> bool:
    for merged in ws.merged_cells.ranges:
        if merged.max_row < min_row or merged.min_row > max_row:
            continue
        if merged.max_col < min_col or merged.min_col > max_col:
            continue
        return True
    return False


def _cell_has_formula(cell) -> bool:
    if cell.data_type == "f":
        return True
    value = cell.value
    return isinstance(value, str) and value.startswith("=")


def inspect_template(path: str) -> TemplateInspection:
    wb = load_workbook(path, data_only=False)
    warnings: list[str] = []

    signature_value = _resolve_defined_name_value(wb, SIGNATURE_NAME)
    signature_ok = signature_value == SIGNATURE_VALUE
    version = "v1" if signature_ok else "unknown"

    blocks: list[ExcelBlock] = []
    for ws in wb.worksheets:
        if ws.protection.sheet:
            warnings.append(f"ERROR: Sheet '{ws.title}' is protected.")

        for table in ws.tables.values():
            table_name = table.displayName or table.name
            if not table_name or not table_name.startswith(BLOCK_PREFIX):
                continue

            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            if _merged_intersects(ws, min_row, max_row, min_col, max_col):
                warnings.append(
                    f"ERROR: Merged cells intersect table '{table_name}' on '{ws.title}'."
                )

            title = ""
            if min_row > 1:
                title_value = ws.cell(row=min_row - 1, column=min_col).value
                if title_value is not None:
                    title = str(title_value)

            col_names = [col.name for col in table.tableColumns]
            replicate_cols = [name for name in col_names if name.startswith("Replicate_")]
            summary_cols = [name for name in ("MEAN", "SEM", "N") if name in col_names]
            metric_col = "Metric" if "Metric" in col_names else ""

            if metric_col:
                col_idx = min_col + col_names.index(metric_col)
                for row in range(min_row + 1, max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    if _cell_has_formula(cell):
                        warnings.append(
                            f"ERROR: Metric column contains formulas in '{table_name}'."
                        )
                        break

            for col_name in replicate_cols:
                col_idx = min_col + col_names.index(col_name)
                for row in range(min_row + 1, max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    if _cell_has_formula(cell):
                        warnings.append(
                            f"ERROR: Replicate column '{col_name}' has formulas in '{table_name}'."
                        )
                        break

            for col_name in summary_cols:
                col_idx = min_col + col_names.index(col_name)
                for row in range(min_row + 1, max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    if cell.value is None:
                        warnings.append(
                            f"ERROR: Summary column '{col_name}' is blank in '{table_name}'."
                        )
                        break
                    if not _cell_has_formula(cell):
                        warnings.append(
                            f"Summary column '{col_name}' lacks formula in '{table_name}'."
                        )
                        break

            blocks.append(
                ExcelBlock(
                    block_id=table_name,
                    sheet_name=ws.title,
                    table_name=table_name,
                    metric_col=metric_col,
                    replicate_cols=replicate_cols,
                    summary_cols=summary_cols,
                    title=title,
                )
            )

    return TemplateInspection(
        signature_ok=signature_ok,
        version=version,
        blocks=blocks,
        warnings=warnings,
    )


def validate_template_or_raise(inspection: TemplateInspection) -> None:
    if not inspection.signature_ok:
        raise ValueError(
            "Template signature missing or invalid. Expected "
            f"{SIGNATURE_NAME} = {SIGNATURE_VALUE}."
        )
    if not inspection.blocks:
        raise ValueError("No VasoAnalyzer data blocks found (VA_Block_* tables).")

    for block in inspection.blocks:
        if block.metric_col != "Metric":
            raise ValueError(
                f"Block '{block.block_id}' is missing required 'Metric' column."
            )
        if "Replicate_1" not in block.replicate_cols:
            raise ValueError(
                f"Block '{block.block_id}' is missing required 'Replicate_1' column."
            )
        missing_summary = [name for name in ("MEAN", "SEM", "N") if name not in block.summary_cols]
        if missing_summary:
            raise ValueError(
                f"Block '{block.block_id}' is missing summary columns: {', '.join(missing_summary)}."
            )

    fatal = [warning for warning in inspection.warnings if warning.startswith("ERROR:")]
    if fatal:
        raise ValueError("\n".join(fatal))


def block_by_id(blocks: Iterable[ExcelBlock], block_id: str) -> ExcelBlock | None:
    for block in blocks:
        if block.block_id == block_id:
            return block
    return None
