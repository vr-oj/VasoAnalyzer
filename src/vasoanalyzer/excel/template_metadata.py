# VasoAnalyzer
# Copyright © 2025 Osvaldo J. Vega Rodríguez
# Licensed under CC BY-NC-SA 4.0 International
# http://creativecommons.org/licenses/by-nc-sa/4.0/

"""
Excel template metadata reader.

This module reads self-describing Excel templates that contain VBA macros
and metadata sheets. It provides a unified API for auto-configuring the
Excel mapping wizard without requiring manual named range setup.

Supported metadata sources (in priority order):
1. VasoMetadata hidden sheet (JSON)
2. Custom document properties
3. Named ranges (fallback for legacy templates)
"""

from __future__ import annotations

import json
import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

__all__ = [
    "TemplateMetadata",
    "EventRowMetadata",
    "DateColumnMetadata",
    "read_template_metadata",
    "has_vaso_metadata",
]

logger = logging.getLogger(__name__)


@dataclass
class EventRowMetadata:
    """Metadata for a single event row in the template."""

    row: int
    label: str
    is_header: bool

    def to_dict(self) -> dict[str, Any]:
        return {"row": self.row, "label": self.label, "is_header": self.is_header}


@dataclass
class DateColumnMetadata:
    """Metadata for a date column in the template."""

    column: int
    letter: str
    value: str | None
    empty_slots: int

    def to_dict(self) -> dict[str, Any]:
        return {
            "column": self.column,
            "letter": self.letter,
            "value": self.value,
            "empty_slots": self.empty_slots,
        }


@dataclass
class TemplateMetadata:
    """Complete metadata describing an Excel template structure."""

    version: str
    template_name: str
    date_row: int
    label_column: int
    event_rows: list[EventRowMetadata] = field(default_factory=list)
    date_columns: list[DateColumnMetadata] = field(default_factory=list)
    source: str = "unknown"  # "vba_metadata", "named_ranges", "inferred"

    # Configuration bounds (optional, from VBA)
    event_rows_start: int | None = None
    event_rows_end: int | None = None
    date_columns_start: int | None = None
    date_columns_end: int | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "version": self.version,
            "template_name": self.template_name,
            "date_row": self.date_row,
            "label_column": self.label_column,
            "event_rows": [row.to_dict() for row in self.event_rows],
            "date_columns": [col.to_dict() for col in self.date_columns],
            "source": self.source,
        }


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def has_vaso_metadata(wb: Workbook) -> bool:
    """
    Check if workbook contains VasoAnalyzer metadata.

    Args:
        wb: openpyxl Workbook instance

    Returns:
        True if metadata is present, False otherwise
    """
    # Check for VasoMetadata sheet
    if "VasoMetadata" in wb.sheetnames:
        return True

    # Check for custom property
    try:
        if hasattr(wb, "custom_doc_props"):
            props = wb.custom_doc_props
            if props and "VasoAnalyzerMetadata" in props:
                return True
    except Exception:
        pass

    return False


def read_template_metadata(
    path: str | Path, wb: Workbook | None = None
) -> TemplateMetadata | None:
    """
    Read template metadata from an Excel file.

    Tries multiple sources in order:
    1. VasoMetadata sheet (hidden sheet with JSON)
    2. Named ranges (VASO_DATES_ROW, VASO_VALUES_BLOCK)
    3. Inference from structure (basic auto-detection)

    Args:
        path: Path to Excel file
        wb: Optional pre-loaded Workbook (if None, will load from path)

    Returns:
        TemplateMetadata if successful, None if no metadata found

    Raises:
        FileNotFoundError: If path doesn't exist
        ValueError: If metadata is malformed
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Template not found: {path}")

    # Load workbook if not provided
    if wb is None:
        try:
            wb = load_workbook(path, data_only=False)
        except Exception as exc:
            logger.error(f"Failed to load workbook {path}: {exc}")
            raise

    ws = wb.active
    if ws is None:
        raise ValueError("Workbook has no active sheet")

    # Try metadata sources in priority order
    metadata = None

    # 1. Try VasoMetadata sheet
    metadata = _read_from_metadata_sheet(wb, ws)
    if metadata:
        metadata.source = "vba_metadata"
        logger.info(f"Read metadata from VasoMetadata sheet: {path}")
        return metadata

    # 2. Try named ranges (legacy)
    metadata = _read_from_named_ranges(wb, ws)
    if metadata:
        metadata.source = "named_ranges"
        logger.info(f"Read metadata from named ranges: {path}")
        return metadata

    # 3. Try inference (basic auto-detection)
    metadata = _infer_from_structure(wb, ws)
    if metadata:
        metadata.source = "inferred"
        logger.warning(f"Inferred metadata from structure: {path}")
        return metadata

    logger.warning(f"No metadata found in template: {path}")
    return None


# ---------------------------------------------------------------------------
# Metadata readers
# ---------------------------------------------------------------------------


def _read_from_metadata_sheet(wb: Workbook, ws: Worksheet) -> TemplateMetadata | None:
    """Read metadata from VasoMetadata hidden sheet."""
    if "VasoMetadata" not in wb.sheetnames:
        return None

    try:
        metadata_sheet = wb["VasoMetadata"]
        # JSON is in cell A4 (see VBA code)
        json_cell = metadata_sheet["A4"]
        if not json_cell.value:
            return None

        data = json.loads(json_cell.value)

        # Parse event rows
        event_rows = [
            EventRowMetadata(
                row=row_data["row"],
                label=row_data["label"],
                is_header=row_data["is_header"],
            )
            for row_data in data.get("event_rows", [])
        ]

        # Parse date columns
        date_columns = [
            DateColumnMetadata(
                column=col_data["column"],
                letter=col_data["letter"],
                value=col_data.get("value"),
                empty_slots=col_data.get("empty_slots", 0),
            )
            for col_data in data.get("date_columns", [])
        ]

        return TemplateMetadata(
            version=data.get("version", "1.0"),
            template_name=data.get("template_name", ""),
            date_row=data["date_row"],
            label_column=data["label_column"],
            event_rows=event_rows,
            date_columns=date_columns,
        )

    except (json.JSONDecodeError, KeyError, TypeError) as exc:
        logger.error(f"Failed to parse VasoMetadata sheet: {exc}")
        return None


def _read_from_named_ranges(wb: Workbook, ws: Worksheet) -> TemplateMetadata | None:
    """Read metadata from named ranges (legacy format)."""
    from openpyxl.utils import range_boundaries

    try:
        # Get named ranges
        date_range = None
        values_range = None

        for name, defn in wb.defined_names.items():
            if name == "VASO_DATES_ROW":
                destinations = list(defn.destinations)
                if destinations:
                    sheet_name, coord = destinations[0]
                    if sheet_name == ws.title:
                        date_range = coord
            elif name == "VASO_VALUES_BLOCK":
                destinations = list(defn.destinations)
                if destinations:
                    sheet_name, coord = destinations[0]
                    if sheet_name == ws.title:
                        values_range = coord

        if not date_range or not values_range:
            return None

        # Parse ranges
        d_min_col, d_min_row, d_max_col, d_max_row = range_boundaries(date_range)
        v_min_col, v_min_row, v_max_col, v_max_row = range_boundaries(values_range)

        # Infer structure from ranges
        metadata = TemplateMetadata(
            version="1.0",
            template_name=wb.properties.title or "",
            date_row=d_min_row,
            label_column=1,  # Assume column A
            event_rows_start=v_min_row,
            event_rows_end=v_max_row,
            date_columns_start=d_min_col,
            date_columns_end=d_max_col,
        )

        # Detect event rows
        metadata.event_rows = _detect_event_rows(
            ws, v_min_row, v_max_row, label_column=1
        )

        # Detect date columns
        metadata.date_columns = _detect_date_columns(
            ws, d_min_row, d_min_col, d_max_col, v_min_row, v_max_row
        )

        return metadata

    except Exception as exc:
        logger.error(f"Failed to read named ranges: {exc}")
        return None


def _infer_from_structure(wb: Workbook, ws: Worksheet) -> TemplateMetadata | None:
    """
    Infer template structure by scanning the worksheet.

    Heuristics:
    - Date row is first row with multiple non-empty cells
    - Event rows start after date row
    - Label column is first column (A)
    """
    try:
        # Find date row (first row with multiple populated cells)
        date_row = None
        for row_idx in range(1, min(10, ws.max_row + 1)):
            non_empty = sum(
                1
                for cell in ws[row_idx]
                if cell.value not in (None, "")
            )
            if non_empty >= 2:
                date_row = row_idx
                break

        if not date_row:
            return None

        # Find event rows (rows with non-empty first cell after date row)
        event_rows_start = date_row + 1
        event_rows_end = min(ws.max_row, event_rows_start + 50)

        metadata = TemplateMetadata(
            version="1.0",
            template_name=wb.properties.title or "Untitled",
            date_row=date_row,
            label_column=1,
            event_rows_start=event_rows_start,
            event_rows_end=event_rows_end,
        )

        # Detect event rows
        metadata.event_rows = _detect_event_rows(
            ws, event_rows_start, event_rows_end, label_column=1
        )

        # Detect date columns (scan date row)
        date_cols_start = 2  # Assume starts at column B
        date_cols_end = min(ws.max_column, 26)  # Up to column Z
        metadata.date_columns = _detect_date_columns(
            ws, date_row, date_cols_start, date_cols_end, event_rows_start, event_rows_end
        )

        return metadata

    except Exception as exc:
        logger.error(f"Failed to infer structure: {exc}")
        return None


# ---------------------------------------------------------------------------
# Detection helpers
# ---------------------------------------------------------------------------


def _detect_event_rows(
    ws: Worksheet, start_row: int, end_row: int, label_column: int = 1
) -> list[EventRowMetadata]:
    """Detect event rows in the worksheet."""
    from openpyxl.styles import PatternFill

    rows = []
    for row_idx in range(start_row, end_row + 1):
        if row_idx > ws.max_row:
            break

        cell = ws.cell(row=row_idx, column=label_column)
        label = str(cell.value).strip() if cell.value not in (None, "") else ""

        if not label:
            continue

        # Check if header (bold + filled)
        is_header = False
        if cell.font and cell.font.bold:
            if cell.fill and isinstance(cell.fill, PatternFill):
                if cell.fill.patternType and cell.fill.patternType != "none":
                    is_header = True

        rows.append(
            EventRowMetadata(row=row_idx, label=label, is_header=is_header)
        )

    return rows


def _detect_date_columns(
    ws: Worksheet,
    date_row: int,
    start_col: int,
    end_col: int,
    event_start_row: int,
    event_end_row: int,
) -> list[DateColumnMetadata]:
    """Detect date columns in the worksheet."""
    from openpyxl.utils import get_column_letter

    columns = []
    for col_idx in range(start_col, end_col + 1):
        if col_idx > ws.max_column:
            break

        cell = ws.cell(row=date_row, column=col_idx)
        value = str(cell.value) if cell.value not in (None, "") else None

        # Count empty slots
        empty_slots = 0
        for row_idx in range(event_start_row, min(event_end_row + 1, ws.max_row + 1)):
            data_cell = ws.cell(row=row_idx, column=col_idx)
            if data_cell.value in (None, ""):
                empty_slots += 1

        columns.append(
            DateColumnMetadata(
                column=col_idx,
                letter=get_column_letter(col_idx),
                value=value,
                empty_slots=empty_slots,
            )
        )

    return columns


# ---------------------------------------------------------------------------
# Optional: VBA macro invocation (requires xlwings)
# ---------------------------------------------------------------------------


def invoke_export_metadata(path: str | Path) -> bool:
    """
    Invoke ExportMetadata VBA macro in the template (if available).

    Requires xlwings library and Excel/LibreOffice installed.

    Args:
        path: Path to .xlsm file

    Returns:
        True if macro executed successfully, False otherwise
    """
    try:
        import xlwings as xw
    except ImportError:
        logger.warning("xlwings not installed, cannot invoke VBA macros")
        return False

    path = Path(path)
    if not path.exists():
        return False

    try:
        # Open workbook with xlwings
        app = xw.App(visible=False)
        wb = app.books.open(str(path.absolute()))

        # Run macro
        macro = wb.macro("ExportMetadata")
        macro()

        # Save and close
        wb.save()
        wb.close()
        app.quit()

        logger.info(f"Successfully invoked ExportMetadata macro: {path}")
        return True

    except Exception as exc:
        logger.error(f"Failed to invoke VBA macro: {exc}")
        return False
