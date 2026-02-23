"""Write session event values into an arbitrary (non-VA-standard) Excel template."""

from __future__ import annotations

import shutil
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from vasoanalyzer.excel.template_metadata import TemplateMetadata


@dataclass
class FlexibleWritePlan:
    """Describes what cells to write when targeting a non-standard template."""

    template_path: str
    output_path: str
    sheet_name: str
    # Each entry: (row_index, col_index, value_or_None)
    writes: list[tuple[int, int, float | None]] = field(default_factory=list)
    # Template labels that had no matching session event
    unmatched_labels: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def build_flexible_write_plan(
    *,
    label_to_event: dict[str, str | None],
    event_label_to_value: dict[str, float],
    metadata: TemplateMetadata,
    target_col: int,
    template_path: str,
    output_path: str,
    sheet_name: str,
) -> FlexibleWritePlan:
    """Build a write plan for a non-standard template.

    Args:
        label_to_event: Maps each template row label → session event label (or None to skip).
        event_label_to_value: Maps session event label → numeric value to write.
        metadata: Detected template metadata (provides event_rows).
        target_col: 1-based column index to write values into.
        template_path: Source template file path.
        output_path: Destination file path.
        sheet_name: Name of the sheet to write into.

    Returns:
        FlexibleWritePlan ready for ``apply_flexible_write_plan``.
    """
    plan = FlexibleWritePlan(
        template_path=template_path,
        output_path=output_path,
        sheet_name=sheet_name,
    )

    for event_row in metadata.event_rows:
        if event_row.is_header:
            continue

        event_label = label_to_event.get(event_row.label)
        if event_label is None:
            plan.unmatched_labels.append(event_row.label)
            continue

        value = event_label_to_value.get(event_label)
        if value is None:
            plan.warnings.append(
                f"No value found for session event '{event_label}' "
                f"(mapped from template row '{event_row.label}')."
            )
            continue

        plan.writes.append((event_row.row, target_col, value))

    if plan.unmatched_labels:
        plan.warnings.append(
            f"{len(plan.unmatched_labels)} template row(s) have no session event assigned "
            f"and will be left blank: {', '.join(plan.unmatched_labels[:5])}"
            + (" …" if len(plan.unmatched_labels) > 5 else "")
        )

    return plan


def apply_flexible_write_plan(plan: FlexibleWritePlan) -> None:
    """Execute a FlexibleWritePlan: copy template to output, write values, save.

    Formula cells are never overwritten — only cells targeted by the plan are
    modified.  The MEAN/SEM/N formulas in the template recalculate automatically
    when Excel opens the saved file.

    Args:
        plan: A plan produced by ``build_flexible_write_plan``.

    Raises:
        FileNotFoundError: If the template file does not exist.
        ValueError: If the plan has no writes.
    """
    template = Path(plan.template_path)
    if not template.exists():
        raise FileNotFoundError(f"Template not found: {template}")

    if not plan.writes:
        raise ValueError("FlexibleWritePlan has no write operations.")

    output = Path(plan.output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    # Copy template → output (preserves formulas, styles, charts)
    shutil.copy2(template, output)

    wb = load_workbook(output, data_only=False)
    if plan.sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{plan.sheet_name}' not found in template. "
            f"Available: {wb.sheetnames}"
        )
    ws = wb[plan.sheet_name]

    for row_idx, col_idx, value in plan.writes:
        ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(output)
