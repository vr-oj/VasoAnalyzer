"""Strict writer for the VasoAnalyzer standard template (v1)."""

from __future__ import annotations

from dataclasses import dataclass
import math
import shutil
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple, get_column_letter, range_boundaries

from .template_v1 import block_by_id, inspect_template, validate_template_or_raise


@dataclass(frozen=True)
class WriteItem:
    metric: str
    value: float
    target_sheet: str
    target_cell: str
    block_id: str
    replicate_col: str


@dataclass(frozen=True)
class WritePlan:
    template_path: str
    output_path: str
    block_id: str
    replicate_col: str
    items: list[WriteItem]
    missing_metrics: list[str]
    warnings: list[str]


def _find_table(ws, table_name: str):
    for table in ws.tables.values():
        name = table.displayName or table.name
        if name == table_name:
            return table
    return None


def _cell_has_formula(cell) -> bool:
    if cell.data_type == "f":
        return True
    value = cell.value
    return isinstance(value, str) and value.startswith("=")


def build_write_plan(
    template_path: str,
    output_path: str,
    block,
    replicate_col: str,
    export_table: pd.DataFrame,
) -> WritePlan:
    if list(export_table.columns) != ["Metric", "Value"]:
        raise ValueError("export_table must have columns exactly: Metric, Value")
    if replicate_col not in block.replicate_cols:
        raise ValueError(
            f"Replicate column '{replicate_col}' not available in block '{block.block_id}'."
        )

    wb = load_workbook(template_path, data_only=False)
    ws = wb[block.sheet_name]
    table = _find_table(ws, block.table_name)
    if table is None:
        raise ValueError(f"Table '{block.table_name}' not found in '{block.sheet_name}'.")

    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    col_names = [col.name for col in table.tableColumns]
    metric_idx = col_names.index("Metric")
    replicate_idx = col_names.index(replicate_col)
    metric_col = min_col + metric_idx
    replicate_col_idx = min_col + replicate_idx

    metrics_to_row: dict[str, int] = {}
    warnings: list[str] = []
    for row in range(min_row + 1, max_row + 1):
        cell = ws.cell(row=row, column=metric_col)
        if cell.value is None:
            continue
        label = str(cell.value).strip()
        if label in metrics_to_row:
            warnings.append(f"Duplicate metric '{label}' in template block '{block.block_id}'.")
            continue
        metrics_to_row[label] = row

    items: list[WriteItem] = []
    missing: list[str] = []
    for _, row in export_table.iterrows():
        metric = str(row["Metric"]).strip()
        if metric not in metrics_to_row:
            missing.append(metric)
            continue
        value = float(row["Value"])
        if math.isnan(value) or math.isinf(value):
            raise ValueError(f"Non-numeric value for metric '{metric}'.")
        target_cell = f"{get_column_letter(replicate_col_idx)}{metrics_to_row[metric]}"
        items.append(
            WriteItem(
                metric=metric,
                value=value,
                target_sheet=block.sheet_name,
                target_cell=target_cell,
                block_id=block.block_id,
                replicate_col=replicate_col,
            )
        )

    return WritePlan(
        template_path=str(template_path),
        output_path=str(output_path),
        block_id=block.block_id,
        replicate_col=replicate_col,
        items=items,
        missing_metrics=missing,
        warnings=warnings,
    )


def validate_write_plan_or_raise(plan: WritePlan) -> None:
    if plan.missing_metrics:
        raise ValueError(
            "Missing metrics in template block: " + ", ".join(plan.missing_metrics)
        )

    inspection = inspect_template(plan.template_path)
    validate_template_or_raise(inspection)
    block = block_by_id(inspection.blocks, plan.block_id)
    if block is None:
        raise ValueError(f"Block '{plan.block_id}' not found in template.")
    if plan.replicate_col not in block.replicate_cols:
        raise ValueError(
            f"Replicate column '{plan.replicate_col}' not available in '{plan.block_id}'."
        )

    wb = load_workbook(plan.template_path, data_only=False)
    ws = wb[block.sheet_name]
    if ws.protection.sheet:
        raise ValueError(f"Sheet '{block.sheet_name}' is protected.")

    table = _find_table(ws, block.table_name)
    if table is None:
        raise ValueError(f"Table '{block.table_name}' not found in '{block.sheet_name}'.")

    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    col_names = [col.name for col in table.tableColumns]
    replicate_idx = col_names.index(plan.replicate_col)
    replicate_col_idx = min_col + replicate_idx

    for item in plan.items:
        row_idx, col_idx = coordinate_to_tuple(item.target_cell)
        if not (min_row + 1 <= row_idx <= max_row):
            raise ValueError(f"Target cell {item.target_cell} is outside table rows.")
        if col_idx != replicate_col_idx:
            raise ValueError(f"Target cell {item.target_cell} is not in replicate column.")
        cell = ws.cell(row=row_idx, column=col_idx)
        if _cell_has_formula(cell):
            raise ValueError(f"Target cell {item.target_cell} contains a formula.")
        for merged in ws.merged_cells.ranges:
            if item.target_cell in merged:
                raise ValueError(f"Target cell {item.target_cell} is merged.")


def apply_write_plan(plan: WritePlan) -> None:
    validate_write_plan_or_raise(plan)

    template_path = Path(plan.template_path)
    output_path = Path(plan.output_path)
    if output_path.resolve() != template_path.resolve():
        output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(template_path, output_path)

    wb = load_workbook(output_path, data_only=False)
    if not plan.items:
        wb.save(output_path)
        return

    ws = wb[plan.items[0].target_sheet]
    for item in plan.items:
        ws[item.target_cell].value = float(item.value)
    wb.save(output_path)
