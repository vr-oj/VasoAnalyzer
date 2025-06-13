from PyQt5.QtWidgets import (
    QDialog, QVBoxLayout, QLabel, QPushButton, QFileDialog, QComboBox,
    QTableWidget, QTableWidgetItem, QHBoxLayout, QMessageBox, QFrame
)
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
import os, sys, subprocess, time

class ExcelMappingDialog(QDialog):
    def __init__(self, parent, event_data):
        super().__init__(parent)
        self.setWindowTitle("Map Events to Excel")
        self.setStyleSheet("""
            QLabel, QTableWidget, QComboBox, QPushButton {
                color: black;
                font-family: Arial;
                font-size: 13px;
            }
            QComboBox {
                background-color: white;
                border: 1px solid #CCCCCC;
                border-radius: 4px;
                padding: 4px;
            }
            QComboBox QAbstractItemView {
                background-color: white;
                selection-background-color: #E6F0FF;
            }
            QPushButton {
                background-color: white;
                border: 1px solid #CCCCCC;
                border-radius: 6px;
                padding: 6px 12px;
            }
            QPushButton:hover {
                background-color: #E6F0FF;
            }
        """)
        self.event_data = event_data
        self.excel_path = None
        self.wb = None
        self.ws = None
        self.current_row = 3
        self.selected_column = None
        self.history = []

        self.setMinimumWidth(460)
        self.layout = QVBoxLayout(self)
        self.layout.setSpacing(12)

        self.instructions = QLabel("<b>Step 1:</b> Select Excel file")
        self.layout.addWidget(self.instructions)

        self.load_button = QPushButton("Load Excel Template")
        self.load_button.clicked.connect(self.load_excel)
        self.layout.addWidget(self.load_button)
        
        self.excel_filename_label = QLabel("")
        self.layout.addWidget(self.excel_filename_label)

        self.layout.addWidget(QLabel("<b>Step 2:</b> Select column to populate:"))
        self.column_selector = QComboBox()
        self.column_selector.addItems([chr(i) for i in range(66, 91)])
        self.column_selector.setEnabled(False)
        self.column_selector.currentTextChanged.connect(self.update_preview_table)
        self.layout.addWidget(self.column_selector)

        self.layout.addSpacing(6)
        self.cell_label = QLabel("Next Excel Cell: N/A")
        self.layout.addWidget(self.cell_label)

        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)
        self.layout.addWidget(line)

        self.event_table = QTableWidget()
        self.event_table.setColumnCount(4)
        self.event_table.setHorizontalHeaderLabels(["EventLabel", "Time (s)", "ID (\u00b5m)", "Frame"])
        self.event_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.event_table.cellClicked.connect(self.map_event_to_excel)
        self.event_table.setMinimumWidth(420)
        self.event_table.horizontalHeader().setStretchLastSection(True)
        self.layout.addWidget(self.event_table)
        self.populate_event_table()

        # Preview table for selected Excel column
        self.preview_table = QTableWidget()
        self.preview_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.preview_table.setMinimumWidth(420)
        self.preview_table.horizontalHeader().setStretchLastSection(True)
        self.layout.addWidget(self.preview_table)

        self.button_layout = QHBoxLayout()
        self.button_layout.addStretch()
        self.skip_button = QPushButton("Skip")
        self.skip_button.clicked.connect(self.skip_cell)
        self.undo_button = QPushButton("Undo Last")
        self.undo_button.clicked.connect(self.undo_last)
        self.done_button = QPushButton("Done")
        self.done_button.clicked.connect(self.finish_and_save)
        self.button_layout.addWidget(self.skip_button)
        self.button_layout.addWidget(self.undo_button)
        self.button_layout.addWidget(self.done_button)
        self.layout.addLayout(self.button_layout)

    def populate_event_table(self):
        self.event_table.setRowCount(len(self.event_data))
        for i, event in enumerate(self.event_data):
            if isinstance(event, dict):
                label = event.get("EventLabel", "")
                time = event.get("Time (s)", "")
                id_val = event.get("ID (µm)", "")
                frame = event.get("Frame", "")
            else:
                label, time, id_val = event
            self.event_table.setItem(i, 0, QTableWidgetItem(str(label)))
            self.event_table.setItem(i, 1, QTableWidgetItem(str(time)))
            self.event_table.setItem(i, 2, QTableWidgetItem(str(id_val)))
            self.event_table.setItem(i, 3, QTableWidgetItem(str(frame)))
        self.event_table.resizeColumnsToContents()
        
        self.event_table.setAlternatingRowColors(True)
        self.event_table.setStyleSheet("""
            QTableWidget {
                alternate-background-color: #F5F5F5;
                background-color: white;
            }
        """)

    def load_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select Excel File", "", "Excel Files (*.xlsx)")
        if path:
            try:
                self.wb = load_workbook(path)
                self.ws = self.wb.active
                self.excel_path = path
                # Ensure the first column remains visible when scrolling
                if not self.ws.freeze_panes:
                    self.ws.freeze_panes = "B1"
                    self.wb.save(path)

                self.column_selector.setEnabled(True)
                self.instructions.setText(
                    "<b>Step 2:</b> File loaded. Now select column and assign values."
                )
                self.update_cell_label()
                self.update_preview_table()
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to load Excel file:\n{e}")
        self.excel_filename_label.setText(f"<i>Loaded:</i> {os.path.basename(path)}")


    def get_current_cell(self):
        col_letter = self.column_selector.currentText()
        return f"{col_letter}{self.current_row}" if col_letter else None

    def update_cell_label(self):
        col_letter = self.column_selector.currentText()
        cell = f"{col_letter}{self.current_row}" if col_letter else "N/A"
        description = ""
        if self.ws:
            try:
                desc_value = str(self.ws[f"A{self.current_row}"].value)
                if desc_value:
                    description = f" \u2192 <i>{desc_value}</i>"
            except:
                pass
        self.cell_label.setText(f"<b>Editing Cell:</b> {cell}{description}")
        self.update_preview_table()

    def update_preview_table(self):
        if not self.ws or not self.column_selector.currentText():
            self.preview_table.setRowCount(0)
            self.preview_table.setColumnCount(0)
            return

        col_letter = self.column_selector.currentText()
        col_idx = column_index_from_string(col_letter)
        column_indices = []
        headers = []

        if col_idx > 1:
            column_indices.append(col_idx - 1)
            headers.append(get_column_letter(col_idx - 1))

        column_indices.append(col_idx)
        headers.append(col_letter)

        column_indices.append(col_idx + 1)
        headers.append(get_column_letter(col_idx + 1))

        start_row = max(1, self.current_row - 4)
        end_row = min(self.ws.max_row, start_row + 8)
        num_rows = end_row - start_row + 1

        self.preview_table.setRowCount(num_rows)
        self.preview_table.setColumnCount(len(headers))
        self.preview_table.setHorizontalHeaderLabels(headers)

        for r, sheet_row in enumerate(range(start_row, end_row + 1)):
            for c, col in enumerate(column_indices):
                value = self.ws.cell(row=sheet_row, column=col).value
                self.preview_table.setItem(r, c, QTableWidgetItem("" if value is None else str(value)))

        self.preview_table.resizeColumnsToContents()

    def map_event_to_excel(self, row, column):
        if not self.ws or not self.column_selector.currentText():
            return
        try:
            col_letter = self.column_selector.currentText()
            value_raw = self.event_table.item(row, 2).text()
            target_cell = f"{col_letter}{self.current_row}"
            try:
                value = float(value_raw)
            except ValueError:
                value = value_raw
            prev_value = self.ws[target_cell].value
            self.history.append((target_cell, prev_value))
            self.ws[target_cell] = value
            self.wb.save(self.excel_path)
            self.current_row += 1
            self.update_cell_label()
            self.update_preview_table()
        except Exception as e:
            QMessageBox.warning(self, "Mapping Error", f"Failed to assign value: {e}")

    def skip_cell(self):
        self.current_row += 1
        self.update_cell_label()
        self.update_preview_table()

    def undo_last(self):
        if not self.history:
            QMessageBox.information(self, "Undo", "Nothing to undo.")
            return
        cell, old_value = self.history.pop()
        self.ws[cell] = old_value
        self.wb.save(self.excel_path)
        self.current_row = int(''.join(filter(str.isdigit, cell)))
        self.update_cell_label()
        self.update_preview_table()

    def finish_and_save(self):
        if self.wb and self.excel_path:
            try:
                self.wb.save(self.excel_path)
                reopen_excel_file_crossplatform(self.excel_path)
                self.accept()
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to save Excel file:\n{e}")

# Auto-update utility

def update_excel_file(excel_path, event_table_data, start_row=3, column_letter="B"):
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        for i, (_, _, id_val, _) in enumerate(event_table_data):  # Now unpacking 4 values
            cell = f"{column_letter}{start_row + i}"
            ws[cell] = id_val
        wb.save(excel_path)
        print(f"🔄 Excel file updated with frame values in column {column_letter}.")
    except Exception as e:
        print(f"❌ Failed to update Excel file:\n{e}")

# Cross-platform file reopening logic

def reopen_excel_file_crossplatform(path):
    try:
        time.sleep(1)
        if sys.platform == "darwin":
            applescript = f'''
            tell application "Microsoft Excel"
                try
                    close (documents whose name is "{os.path.basename(path)}") saving yes
                end try
                open POSIX file "{path}"
                activate
            end tell
            '''

            subprocess.call(["osascript", "-e", applescript])
        elif sys.platform == "win32":
            os.startfile(path)
        elif sys.platform.startswith("linux"):
            subprocess.call(["xdg-open", path])
    except Exception as e:
        print(f"\u26a0\ufe0f Could not reopen Excel file:\n{e}")
